from django.shortcuts import get_object_or_404
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from .models import Framework, Policy, SubPolicy, FrameworkVersion, PolicyVersion, PolicyApproval, Users
from .serializers import FrameworkSerializer, PolicySerializer, SubPolicySerializer, PolicyApprovalSerializer, UserSerializer   
from django.db import transaction
import traceback
import sys
import datetime
import re
from django.utils import timezone
from datetime import timedelta
from django.db.models import Count, Avg, Case, When, Value, FloatField, F
from django.db.models.functions import Coalesce, Cast
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework import status
from .models import Users
import json
import logging

logger = logging.getLogger(__name__)

# Framework CRUD operations

"""
@api GET /api/frameworks/
Returns all frameworks with Status='Approved' and ActiveInactive='Active'.
Filtered by the serializer to include only policies with Status='Approved' and ActiveInactive='Active',
and subpolicies with Status='Approved'.

@api POST /api/frameworks/
Creates a new framework with associated policies and subpolicies.
New frameworks are created with Status='Under Review' and ActiveInactive='Inactive' by default.
CurrentVersion defaults to 1.0 if not provided.

Example payload:
{
  "FrameworkName": "ISO 27001",
  "FrameworkDescription": "Information Security Management System",
  "EffectiveDate": "2023-10-01",
  "CreatedByName": "John Doe",
  "CreatedByDate": "2023-09-15",
  "Category": "Information Security and Compliance",
  "DocURL": "https://example.com/iso27001",
  "Identifier": "ISO-27001",
  "StartDate": "2023-10-01",
  "EndDate": "2025-10-01",
  "policies": [
    {
      "PolicyName": "Access Control Policy",
      "PolicyDescription": "Guidelines for access control management",
      "StartDate": "2023-10-01",
      "Department": "IT",
      "Applicability": "All Employees",
      "Scope": "All IT systems",
      "Objective": "Ensure proper access control",
      "Identifier": "ACP-001",
      "subpolicies": [
        {
          "SubPolicyName": "Password Management",
          "Identifier": "PWD-001",
          "Description": "Password requirements and management",
          "PermanentTemporary": "Permanent",
          "Control": "Use strong passwords with at least 12 characters"
        }
      ]
    }
  ]
}
"""
@api_view(['GET', 'POST'])
@permission_classes([AllowAny])
def framework_list(request):
    if request.method == 'GET':
        frameworks = Framework.objects.filter(Status='Approved', ActiveInactive='Active')
        serializer = FrameworkSerializer(frameworks, many=True)
        return Response(serializer.data)
 
    elif request.method == 'POST':
        try:
            with transaction.atomic():
                # Prepare incoming data
                data = request.data.copy()
 
                # Set default values if not provided
                data.setdefault('Status', 'Under Review')
                data.setdefault('ActiveInactive', 'Inactive')
               
                # Always set CreatedByDate to current date
                data['CreatedByDate'] = datetime.date.today()
 
                # Set version to 1.0 for all new frameworks
                new_version = 1.0
 
                # Create Framework
                framework_serializer = FrameworkSerializer(data=data)
                if not framework_serializer.is_valid():
                    return Response(framework_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
 
                framework = framework_serializer.save()
                framework.CurrentVersion = new_version
                framework.save()
 
                # Create FrameworkVersion
                framework_version = FrameworkVersion(
                    FrameworkId=framework,
                    Version=framework.CurrentVersion,
                    FrameworkName=framework.FrameworkName,
                    CreatedBy=framework.CreatedByName,
                    CreatedDate=datetime.date.today(),  # Always use current date
                    PreviousVersionId=None
                )
                framework_version.save()
 
                # Handle Policies if provided
                policies_data = request.data.get('policies', [])
                created_policies_count = 0
                created_subpolicies_count = 0
               
                for policy_data in policies_data:
                    policy_data = policy_data.copy()
                    policy_data['FrameworkId'] = framework.FrameworkId
                    policy_data['CurrentVersion'] = framework.CurrentVersion
                    policy_data.setdefault('Status', 'Under Review')
                    policy_data.setdefault('ActiveInactive', 'Inactive')
                    policy_data.setdefault('CreatedByName', framework.CreatedByName)
                    policy_data['CreatedByDate'] = datetime.date.today()  # Always use current date
                   
                    # Get reviewer's name if reviewer ID is provided
                    reviewer_id = policy_data.get('Reviewer')
                    if reviewer_id:
                        reviewer_obj = Users.objects.filter(UserId=reviewer_id).first()
                        if reviewer_obj:
                            # Store the reviewer's name in the policy
                            policy_data['Reviewer'] = reviewer_obj.UserName
 
                    policy_serializer = PolicySerializer(data=policy_data)
                    if not policy_serializer.is_valid():
                        return Response(policy_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
 
                    policy = policy_serializer.save()
                    created_policies_count += 1
 
                    policy_version = PolicyVersion(
                        PolicyId=policy,
                        Version=policy.CurrentVersion,
                        PolicyName=policy.PolicyName,
                        CreatedBy=policy.CreatedByName,
                        CreatedDate=datetime.date.today(),  # Always use current date
                        PreviousVersionId=None
                    )
                    policy_version.save()
                   
                    # Handle SubPolicies if provided
                    subpolicies_data = policy_data.get('subpolicies', [])
                    for subpolicy_data in subpolicies_data:
                        subpolicy_data = subpolicy_data.copy()
                        subpolicy_data['PolicyId'] = policy.PolicyId
                        subpolicy_data.setdefault('Status', 'Under Review')
                        subpolicy_data.setdefault('CreatedByName', policy.CreatedByName)
                        subpolicy_data['CreatedByDate'] = datetime.date.today()  # Always use current date
 
                        subpolicy_serializer = SubPolicySerializer(data=subpolicy_data)
                        if not subpolicy_serializer.is_valid():
                            return Response(subpolicy_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
                        subpolicy_serializer.save()
                        created_subpolicies_count += 1
 
                # Create a detailed success message
                message = f'Framework "{framework.FrameworkName}" created successfully'
                if created_policies_count > 0:
                    message += f' with {created_policies_count} policies'
                    if created_subpolicies_count > 0:
                        message += f' and {created_subpolicies_count} subpolicies'
                message += '!'
               
                return Response({
                    'message': message,
                    'FrameworkId': framework.FrameworkId,
                    'Version': framework.CurrentVersion
                }, status=status.HTTP_201_CREATED)
 
        except Exception as e:
            return Response({
                'error': 'Error creating framework',
                'details': {
                    'message': str(e),
                    'traceback': traceback.format_exc()
                }
            }, status=status.HTTP_400_BAD_REQUEST)
 
 

"""
@api GET /api/frameworks/{pk}/
Returns a specific framework by ID if it has Status='Approved' and ActiveInactive='Active'.

@api PUT /api/frameworks/{pk}/
Updates an existing framework. Only frameworks with Status='Approved' and ActiveInactive='Active' can be updated.

Example payload:
{
  "FrameworkName": "ISO 27001:2022",
  "FrameworkDescription": "Updated Information Security Management System",
  "Category": "Information Security",
  "DocURL": "https://example.com/iso27001-2022",
  "EndDate": "2026-10-01"
}

@api DELETE /api/frameworks/{pk}/
Soft-deletes a framework by setting ActiveInactive='Inactive'.
Also marks all related policies as inactive and all related subpolicies with Status='Inactive'.
"""
@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([AllowAny])
def framework_detail(request, pk):
    framework = get_object_or_404(Framework, FrameworkId=pk)
    
    if request.method == 'GET':
        # Only return details if framework is Approved and Active
        if framework.Status != 'Approved' or framework.ActiveInactive != 'Active':
            return Response({'error': 'Framework is not approved or active'}, status=status.HTTP_403_FORBIDDEN)
        
        # Get all active and approved policies for this framework
        policies = Policy.objects.filter(
            FrameworkId=framework,
            Status='Approved',
            ActiveInactive='Active'
        )
        
        # Get all subpolicies for these policies
        policy_data = []
        for policy in policies:
            policy_dict = {
                'PolicyId': policy.PolicyId,
                'PolicyName': policy.PolicyName,
                'PolicyDescription': policy.PolicyDescription,
                'CurrentVersion': policy.CurrentVersion,
                'StartDate': policy.StartDate,
                'EndDate': policy.EndDate,
                'Department': policy.Department,
                'CreatedByName': policy.CreatedByName,
                'CreatedByDate': policy.CreatedByDate,
                'Applicability': policy.Applicability,
                'DocURL': policy.DocURL,
                'Scope': policy.Scope,
                'Objective': policy.Objective,
                'Identifier': policy.Identifier,
                'PermanentTemporary': policy.PermanentTemporary,
                'subpolicies': []
            }
            
            # Get all subpolicies for this policy
            subpolicies = SubPolicy.objects.filter(PolicyId=policy)
            for subpolicy in subpolicies:
                subpolicy_dict = {
                    'SubPolicyId': subpolicy.SubPolicyId,
                    'SubPolicyName': subpolicy.SubPolicyName,
                    'CreatedByName': subpolicy.CreatedByName,
                    'CreatedByDate': subpolicy.CreatedByDate,
                    'Identifier': subpolicy.Identifier,
                    'Description': subpolicy.Description,
                    'Status': subpolicy.Status,
                    'PermanentTemporary': subpolicy.PermanentTemporary,
                    'Control': subpolicy.Control
                }
                policy_dict['subpolicies'].append(subpolicy_dict)
            
            policy_data.append(policy_dict)
        
        # Create response data
        response_data = {
            'FrameworkId': framework.FrameworkId,
            'FrameworkName': framework.FrameworkName,
            'CurrentVersion': framework.CurrentVersion,
            'FrameworkDescription': framework.FrameworkDescription,
            'EffectiveDate': framework.EffectiveDate,
            'CreatedByName': framework.CreatedByName,
            'CreatedByDate': framework.CreatedByDate,
            'Category': framework.Category,
            'DocURL': framework.DocURL,
            'Identifier': framework.Identifier,
            'StartDate': framework.StartDate,
            'EndDate': framework.EndDate,
            'Status': framework.Status,
            'ActiveInactive': framework.ActiveInactive,
            'policies': policy_data
        }
        
        return Response(response_data)
    
    elif request.method == 'PUT':
        # Check if framework is approved and active before allowing update
        if framework.Status != 'Approved' or framework.ActiveInactive != 'Active':
            return Response({'error': 'Only approved and active frameworks can be updated'}, status=status.HTTP_403_FORBIDDEN)
        
        try:
            with transaction.atomic():
                serializer = FrameworkSerializer(framework, data=request.data, partial=True)
                if serializer.is_valid():
                    serializer.save()
                    return Response({
                        'message': 'Framework updated successfully',
                        'FrameworkId': framework.FrameworkId,
                        'CurrentVersion': framework.CurrentVersion
                    })
                else:
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            error_info = {
                'error': str(e),
                'traceback': traceback.format_exc()
            }
            return Response({'error': 'Error updating framework', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'DELETE':
        try:
            with transaction.atomic():
                # Instead of deleting, set ActiveInactive to 'Inactive'
                framework.ActiveInactive = 'Inactive'
                framework.save()
                
                # Set all related policies to inactive
                policies = Policy.objects.filter(FrameworkId=framework)
                for policy in policies:
                    policy.ActiveInactive = 'Inactive'
                    policy.save()
                    
                    # Update Status of subpolicies since they don't have ActiveInactive field
                    subpolicies = SubPolicy.objects.filter(PolicyId=policy)
                    for subpolicy in subpolicies:
                        subpolicy.Status = 'Inactive'
                        subpolicy.save()
                
                return Response({'message': 'Framework and related policies marked as inactive'}, status=status.HTTP_200_OK)
        except Exception as e:
            error_info = {
                'error': str(e),
                'traceback': traceback.format_exc()
            }
            return Response({'error': 'Error marking framework as inactive', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)

# Policy CRUD operations

"""
@api GET /api/policies/{pk}/
Returns a specific policy by ID if it has Status='Approved' and ActiveInactive='Active',
and its parent framework has Status='Approved' and ActiveInactive='Active'.

@api PUT /api/policies/{pk}/
Updates an existing policy. Only policies with Status='Approved' and ActiveInactive='Active'
whose parent framework is also Approved and Active can be updated.

Example payload:
{
  "PolicyName": "Updated Access Control Policy",
  "PolicyDescription": "Enhanced guidelines for access control management with additional security measures",
  "StartDate": "2023-12-01",
  "EndDate": "2025-12-01",
  "Department": "IT,Security",
  "Scope": "All IT systems and cloud services",
  "Objective": "Ensure proper access control with improved security"
}

@api DELETE /api/policies/{pk}/
Soft-deletes a policy by setting ActiveInactive='Inactive'.
Also marks all related subpolicies with Status='Inactive'.
"""
@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([AllowAny])
def policy_detail(request, pk):
    policy = get_object_or_404(Policy, PolicyId=pk)
    
    if request.method == 'GET':
        # Only return details if policy is Approved and Active
        if policy.Status != 'Approved' or policy.ActiveInactive != 'Active':
            return Response({'error': 'Policy is not approved or active'}, status=status.HTTP_403_FORBIDDEN)
        
        # Get framework to check if it's approved and active
        framework = policy.FrameworkId
        if framework.Status != 'Approved' or framework.ActiveInactive != 'Active':
            return Response({'error': 'Framework is not approved or active'}, status=status.HTTP_403_FORBIDDEN)
        
        serializer = PolicySerializer(policy)
        return Response(serializer.data)
    
    elif request.method == 'PUT':
        # Check if policy is approved and active before allowing update
        if policy.Status != 'Approved' or policy.ActiveInactive != 'Active':
            return Response({'error': 'Only approved and active policies can be updated'}, status=status.HTTP_403_FORBIDDEN)
        
        # Check if framework is approved and active
        framework = policy.FrameworkId
        if framework.Status != 'Approved' or framework.ActiveInactive != 'Active':
            return Response({'error': 'Framework is not approved or active'}, status=status.HTTP_403_FORBIDDEN)
        
        try:
            with transaction.atomic():
                # Add status and ActiveInactive to request data
                update_data = request.data.copy()
                update_data['Status'] = 'Under Review'
                update_data['ActiveInactive'] = 'Inactive'
                
                serializer = PolicySerializer(policy, data=update_data, partial=True)
                if serializer.is_valid():
                    serializer.save()
                    return Response({
                        'message': 'Policy updated successfully and set to Under Review',
                        'PolicyId': policy.PolicyId,
                        'CurrentVersion': policy.CurrentVersion,
                        'Status': 'Under Review',
                        'ActiveInactive': 'Inactive'
                    })
                else:
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            error_info = {
                'error': str(e),
                'traceback': traceback.format_exc()
            }
            return Response({'error': 'Error updating policy', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'DELETE':
        try:
            with transaction.atomic():
                # Instead of deleting, set ActiveInactive to 'Inactive'
                policy.ActiveInactive = 'Inactive'
                policy.save()
                
                # Update Status of subpolicies since they don't have ActiveInactive field
                subpolicies = SubPolicy.objects.filter(PolicyId=policy)
                for subpolicy in subpolicies:
                    subpolicy.Status = 'Inactive'
                    subpolicy.save()
                
                return Response({'message': 'Policy and related subpolicies marked as inactive'}, status=status.HTTP_200_OK)
        except Exception as e:
            error_info = {
                'error': str(e),
                'traceback': traceback.format_exc()
            }
            return Response({'error': 'Error marking policy as inactive', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)

"""
@api POST /api/frameworks/{framework_id}/policies/
Adds a new policy to an existing framework.
New policies are created with Status='Under Review' and ActiveInactive='Inactive' by default.
CurrentVersion defaults to 1.0 if not provided.

Example payload:
{
  "PolicyName": "Data Classification Policy",
  "PolicyDescription": "Guidelines for data classification and handling",
  "StartDate": "2023-10-01",
  "Department": "IT,Legal",
  "Applicability": "All Employees",
  "Scope": "All company data",
  "Objective": "Ensure proper data classification and handling",
  "Identifier": "DCP-001",
  "subpolicies": [
    {
      "SubPolicyName": "Confidential Data Handling",
      "Identifier": "CDH-001",
      "Description": "Guidelines for handling confidential data",
      "PermanentTemporary": "Permanent",
      "Control": "Encrypt all confidential data at rest and in transit"
    }
  ]
}
"""
@api_view(['POST'])
@permission_classes([AllowAny])
def add_policy_to_framework(request, framework_id):
    framework = get_object_or_404(Framework, FrameworkId=framework_id)
   
    try:
        with transaction.atomic():
            # Set framework ID and default values in the request data
            policy_data = request.data.copy()
            policy_data['FrameworkId'] = framework.FrameworkId
            policy_data['CurrentVersion'] = framework.CurrentVersion  # Use framework's version
            
            # Set default values if not provided
            if 'Status' not in policy_data:
                policy_data['Status'] = 'Under Review'
            if 'ActiveInactive' not in policy_data:
                policy_data['ActiveInactive'] = 'Inactive'
            if 'CreatedByName' not in policy_data or not policy_data['CreatedByName']:
                policy_data['CreatedByName'] = framework.CreatedByName
            if 'CreatedByDate' not in policy_data:
                policy_data['CreatedByDate'] = datetime.date.today()
            if 'Reviewer' not in policy_data:
                policy_data['Reviewer'] = None
           
            print("DEBUG: Policy data before serialization:", policy_data)
            policy_serializer = PolicySerializer(data=policy_data)
            print("DEBUG: validating policy serializer")
            if not policy_serializer.is_valid():
                print("Policy serializer errors:", policy_serializer.errors)
                return Response({
                    'error': 'Policy validation failed',
                    'details': policy_serializer.errors
                }, status=status.HTTP_400_BAD_REQUEST)
            print("DEBUG: serializer is valid")
 
            policy = policy_serializer.save()
 
            # Get reviewer ID directly from the request data
            reviewer_id = policy_data.get('Reviewer')  # This should be a UserId (number)
           
            # Get reviewer's name for the Policy table
            reviewer_name = None
            if reviewer_id:
                try:
                    reviewer_id = int(reviewer_id)  # Ensure reviewer_id is an integer
                    reviewer_obj = Users.objects.filter(UserId=reviewer_id).first()
                    if reviewer_obj:
                        reviewer_name = reviewer_obj.UserName
                        # Store reviewer name in the policy object
                        policy.Reviewer = reviewer_name
                        policy.save()
                except (ValueError, TypeError):
                    print(f"Warning: Invalid reviewer ID format: {reviewer_id}")
 
            # Get user id from CreatedByName
            created_by_name = policy_data.get('CreatedByName')
            user_obj = Users.objects.filter(UserName=created_by_name).first()
            user_id = user_obj.UserId if user_obj else None
 
            if user_id is None:
                print(f"Warning: CreatedBy user not found for: {created_by_name}")
            if reviewer_id is None:
                print("Warning: Reviewer id missing in request data")

            try:
                print("Creating PolicyVersion with:", {
                    "PolicyId": policy.PolicyId,
                    "Version": policy.CurrentVersion,
                    "PolicyName": policy.PolicyName,
                    "CreatedBy": policy.CreatedByName,
                    "CreatedDate": policy.CreatedByDate,
                    "PreviousVersionId": None
                })
 
                policy_version = PolicyVersion(
                    PolicyId=policy,
                    Version=policy.CurrentVersion,
                    PolicyName=policy.PolicyName,
                    CreatedBy=policy.CreatedByName,
                    CreatedDate=policy.CreatedByDate,
                    PreviousVersionId=None
                )
                policy_version.save()
            except Exception as e:
                print("Error creating PolicyVersion:", str(e))
                raise
 
           
            # Create subpolicies if provided
            subpolicies_data = request.data.get('subpolicies', [])
            created_subpolicies_count = 0
           
            for subpolicy_data in subpolicies_data:
                # Set policy ID and default values
                subpolicy_data = subpolicy_data.copy() if isinstance(subpolicy_data, dict) else {}
                subpolicy_data['PolicyId'] = policy.PolicyId
                if 'CreatedByName' not in subpolicy_data or not subpolicy_data['CreatedByName']:
                    subpolicy_data['CreatedByName'] = policy.CreatedByName
                if 'CreatedByDate' not in subpolicy_data:
                    subpolicy_data['CreatedByDate'] = datetime.date.today()
                if 'Status' not in subpolicy_data:
                    subpolicy_data['Status'] = 'Under Review'
                if 'PermanentTemporary' not in subpolicy_data:
                    subpolicy_data['PermanentTemporary'] = 'Permanent'
               
                print("DEBUG: SubPolicy data before serialization:", subpolicy_data)
                subpolicy_serializer = SubPolicySerializer(data=subpolicy_data)
                if not subpolicy_serializer.is_valid():
                    print("SubPolicy serializer errors:", subpolicy_serializer.errors)
                    return Response({
                        'error': 'SubPolicy validation failed',
                        'details': subpolicy_serializer.errors
                    }, status=status.HTTP_400_BAD_REQUEST)
                subpolicy_serializer.save()
                created_subpolicies_count += 1
               
            # Create a detailed success message
            message = 'Policy added to framework successfully'
            if created_subpolicies_count > 0:
                message += f' with {created_subpolicies_count} subpolicies'
            message += '!'
           
            return Response({
                'message': message,
                'PolicyId': policy.PolicyId,
                'FrameworkId': framework.FrameworkId,
                'Version': policy.CurrentVersion
            }, status=status.HTTP_201_CREATED)
    except Exception as e:
        error_info = {
            'error': str(e),
            'traceback': traceback.format_exc()
        }
        print("DEBUG: Error details:", error_info)
        return Response({
            'error': 'Error adding policy to framework',
            'details': error_info
        }, status=status.HTTP_400_BAD_REQUEST)

"""
@api POST /api/policies/{policy_id}/subpolicies/
Adds a new subpolicy to an existing policy.
New subpolicies are created with Status='Under Review' by default.

Example payload:
{
  "SubPolicyName": "Multi-Factor Authentication",
  "Identifier": "MFA-001",
  "Description": "Requirements for multi-factor authentication",
  "PermanentTemporary": "Permanent",
  "Control": "Implement MFA for all admin access and sensitive operations"
}
"""
@api_view(['POST'])
@permission_classes([AllowAny])
def add_policy_to_framework(request, framework_id):
    framework = get_object_or_404(Framework, FrameworkId=framework_id)
   
    try:
        with transaction.atomic():
            # Set framework ID and default values in the request data
            policy_data = request.data.copy()
            policy_data['FrameworkId'] = framework.FrameworkId
            policy_data['CurrentVersion'] = framework.CurrentVersion  # Use framework's version
            if 'Status' not in policy_data:
                policy_data['Status'] = 'Under Review'
            if 'ActiveInactive' not in policy_data:
                policy_data['ActiveInactive'] = 'Inactive'
            if 'CreatedByName' not in policy_data:
                policy_data['CreatedByName'] = framework.CreatedByName
            if 'CreatedByDate' not in policy_data:
                policy_data['CreatedByDate'] = datetime.date.today()
           
            policy_serializer = PolicySerializer(data=policy_data)
            print("DEBUG: validating policy serializer")
            if not policy_serializer.is_valid():
                print("Policy serializer errors:", policy_serializer.errors)
                return Response(policy_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            print("DEBUG: serializer is valid")
 
            policy = policy_serializer.save()
 
            # Get reviewer ID directly from the request data
            reviewer_id = policy_data.get('Reviewer')  # Changed from request.data to policy_data
           
            # Get reviewer's name for the Policy table
            reviewer_name = None
            if reviewer_id:
                reviewer_obj = Users.objects.filter(UserId=reviewer_id).first()
                if reviewer_obj:
                    reviewer_name = reviewer_obj.UserName
                    # Store reviewer name in the policy object
                    policy.Reviewer = reviewer_name
                    policy.save()
 
            # Get user id from CreatedByName
            created_by_name = policy_data.get('CreatedByName')
            user_obj = Users.objects.filter(UserName=created_by_name).first()
            user_id = user_obj.UserId if user_obj else None
 
            if user_id is None:
                print(f"Warning: CreatedBy user not found for: {created_by_name}")
            if reviewer_id is None:
                print("Warning: Reviewer id missing in request data")
 
            # No policy approval logic here - removed completely
 
            try:
                print("Creating PolicyVersion with:", {
                    "PolicyId": policy.PolicyId,
                    "Version": policy.CurrentVersion,
                    "PolicyName": policy.PolicyName,
                    "CreatedBy": policy.CreatedByName,
                    "CreatedDate": policy.CreatedByDate,
                    "PreviousVersionId": None
                })
 
                policy_version = PolicyVersion(
                    PolicyId=policy,
                    Version=policy.CurrentVersion,
                    PolicyName=policy.PolicyName,
                    CreatedBy=policy.CreatedByName,
                    CreatedDate=policy.CreatedByDate,
                    PreviousVersionId=None
                )
                policy_version.save()
            except Exception as e:
                print("Error creating PolicyVersion:", str(e))
                raise
 
            # Create subpolicies if provided
            subpolicies_data = request.data.get('subpolicies', [])
            created_subpolicies_count = 0
           
            for subpolicy_data in subpolicies_data:
                # Set policy ID and default values
                subpolicy_data = subpolicy_data.copy() if isinstance(subpolicy_data, dict) else {}
                subpolicy_data['PolicyId'] = policy.PolicyId
                if 'CreatedByName' not in subpolicy_data:
                    subpolicy_data['CreatedByName'] = policy.CreatedByName
                if 'CreatedByDate' not in subpolicy_data:
                    subpolicy_data['CreatedByDate'] = policy.CreatedByDate
                if 'Status' not in subpolicy_data:
                    subpolicy_data['Status'] = 'Under Review'
               
                subpolicy_serializer = SubPolicySerializer(data=subpolicy_data)
                if not subpolicy_serializer.is_valid():
                    print("SubPolicy serializer errors:", subpolicy_serializer.errors)  # Add this debug
                    return Response(subpolicy_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
                subpolicy_serializer.save()
                created_subpolicies_count += 1
               
            # Create a detailed success message
            message = 'Policy added to framework successfully'
            if created_subpolicies_count > 0:
                message += f' with {created_subpolicies_count} subpolicies'
            message += '!'
           
            return Response({
                'message': message,
                'PolicyId': policy.PolicyId,
                'FrameworkId': framework.FrameworkId,
                'Version': policy.CurrentVersion
            }, status=status.HTTP_201_CREATED)
    except Exception as e:
        error_info = {
            'error': str(e),
            'traceback': traceback.format_exc()
        }
        return Response({'error': 'Error adding policy to framework', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([AllowAny])
def list_policy_approvals_for_reviewer(request):
    # For now, reviewer_id is hardcoded as 2
    reviewer_id = 2
   
    # Get the latest version of each policy by identifier
    unique_identifiers = PolicyApproval.objects.values('Identifier').distinct()
    latest_approvals = []
   
    for identifier_obj in unique_identifiers:
        identifier = identifier_obj['Identifier']
        # Find the latest approval record for this identifier
        latest = PolicyApproval.objects.filter(
            Identifier=identifier,
            ReviewerId=reviewer_id
        ).order_by('-ApprovalId').first()
       
        if latest:
            latest_approvals.append(latest)
   
    # Serialize the data
    data = [
        {
            "ApprovalId": a.ApprovalId,
            "Identifier": a.Identifier,
            "ExtractedData": a.ExtractedData,
            "UserId": a.UserId,
            "ReviewerId": a.ReviewerId,
            "ApprovedNot": a.ApprovedNot,
            "Version": a.Version
        }
        for a in latest_approvals
    ]
   
    return Response(data)
 
@api_view(['PUT'])
@permission_classes([AllowAny])
def update_policy_approval(request, approval_id):
    try:
        # Get the original approval
        approval = PolicyApproval.objects.get(ApprovalId=approval_id)
       
        # Create a new approval object instead of updating
        new_approval = PolicyApproval()
        new_approval.Identifier = approval.Identifier
        new_approval.ExtractedData = request.data.get('ExtractedData', approval.ExtractedData)
        new_approval.UserId = approval.UserId
        new_approval.ReviewerId = approval.ReviewerId
        new_approval.ApprovedNot = request.data.get('ApprovedNot', approval.ApprovedNot)
       
        # Determine version prefix based on who made the change
        # For reviewers (assuming ReviewerId is the one making changes in this endpoint)
        prefix = 'r'
       
        # Get the latest version with this prefix for this identifier
        latest_version = PolicyApproval.objects.filter(
            Identifier=approval.Identifier,
            Version__startswith=prefix
        ).order_by('-Version').first()
       
        if latest_version and latest_version.Version:
            # Extract number and increment
            try:
                version_num = int(latest_version.Version[1:])
                new_approval.Version = f"{prefix}{version_num + 1}"
            except ValueError:
                new_approval.Version = f"{prefix}1"
        else:
            new_approval.Version = f"{prefix}1"
       
        new_approval.save()
       
        return Response({
            'message': 'Policy approval updated successfully',
            'ApprovalId': new_approval.ApprovalId,
            'Version': new_approval.Version
        })
    except PolicyApproval.DoesNotExist:
        return Response({'error': 'Policy approval not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
 
@api_view(['PUT'])
@permission_classes([AllowAny])
def resubmit_policy_approval(request, approval_id):
    try:
        # Get the original approval
        approval = PolicyApproval.objects.get(ApprovalId=approval_id)
       
        # Validate data
        extracted_data = request.data.get('ExtractedData')
        if not extracted_data:
            return Response({'error': 'ExtractedData is required'}, status=status.HTTP_400_BAD_REQUEST)
       
        # Print debug info
        print(f"Resubmitting policy with ID: {approval_id}, Identifier: {approval.Identifier}")
       
        # Get all versions for this identifier with 'u' prefix
        all_versions = PolicyApproval.objects.filter(Identifier=approval.Identifier)
       
        # Find the highest 'u' version number
        highest_u_version = 0
        for pa in all_versions:
            if pa.Version and pa.Version.startswith('u') and len(pa.Version) > 1:
                try:
                    version_num = int(pa.Version[1:])
                    if version_num > highest_u_version:
                        highest_u_version = version_num
                except ValueError:
                    continue
       
        # Set the new version
        new_version = f"u{highest_u_version + 1}"
        print(f"Setting new version: {new_version}")
       
        # Create a new approval object manually
        new_approval = PolicyApproval(
            Identifier=approval.Identifier,
            ExtractedData=extracted_data,
            UserId=approval.UserId,
            ReviewerId=approval.ReviewerId,
            ApprovedNot=None,  # Reset approval status
            Version=new_version
        )
       
        # Reset subpolicy approvals
        if 'subpolicies' in extracted_data and isinstance(extracted_data['subpolicies'], list):
            for subpolicy in extracted_data['subpolicies']:
                if subpolicy.get('approval', {}).get('approved') == False:
                    subpolicy['approval'] = {
                        'approved': None,
                        'remarks': ''
                    }
       
        # Save the new record
        new_approval.save()
        print(f"Saved new approval with ID: {new_approval.ApprovalId}, Version: {new_approval.Version}")
       
        return Response({
            'message': 'Policy resubmitted for review successfully',
            'ApprovalId': new_approval.ApprovalId,
            'Version': new_version
        })
       
    except PolicyApproval.DoesNotExist:
        return Response({'error': 'Policy approval not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        print("Error in resubmit_policy_approval:", str(e))
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
 
@api_view(['GET'])
@permission_classes([AllowAny])
def list_rejected_policy_approvals_for_user(request, user_id):
    # Filter policies by ReviewerId (not UserId) since we want reviewer's view
    approvals = PolicyApproval.objects.filter(ReviewerId=user_id)
   
    # Group by Identifier to find the latest for each policy/compliance
    identifier_latest = {}
    for approval in approvals:
        identifier = approval.Identifier
        if identifier not in identifier_latest or approval.ApprovalId > identifier_latest[identifier].ApprovalId:
            identifier_latest[identifier] = approval
   
    result = []
    for identifier, approval in identifier_latest.items():
        extracted = approval.ExtractedData
       
        # Check if this is a compliance item
        is_compliance = extracted.get('type') == 'compliance'
       
        # Check if main policy/compliance is rejected
        main_rejected = approval.ApprovedNot is False
       
        if is_compliance:
            # For compliance items
            item_rejected = extracted.get('compliance_approval', {}).get('approved') is False
           
            if main_rejected or item_rejected:
                result.append({
                    "ApprovalId": approval.ApprovalId,
                    "Identifier": approval.Identifier,
                    "ExtractedData": approval.ExtractedData,
                    "UserId": approval.UserId,
                    "ReviewerId": approval.ReviewerId,
                    "ApprovedNot": approval.ApprovedNot,
                    "main_item_rejected": main_rejected,
                    "is_compliance": True,
                    "rejection_reason": extracted.get('compliance_approval', {}).get('remarks', "")
                })
        else:
            # For policy items (existing logic)
            main_policy_rejected = main_rejected or (
                extracted.get('policy_approval', {}).get('approved') is False
            )
            # Check for rejected subpolicies
            rejected_subpolicies = []
            for sub in extracted.get('subpolicies', []):
                if sub.get('approval', {}).get('approved') is False:
                    rejected_subpolicies.append({
                        "Identifier": sub.get("Identifier"),
                        "SubPolicyName": sub.get("SubPolicyName"),
                        "Description": sub.get("Description"),
                        "Control": sub.get("Control"),
                        "remarks": sub.get("approval", {}).get("remarks", "")
                    })
            # Only add if main policy or any subpolicy is rejected
            if main_policy_rejected or rejected_subpolicies:
                result.append({
                    "ApprovalId": approval.ApprovalId,
                    "Identifier": approval.Identifier,
                    "ExtractedData": approval.ExtractedData,
                    "UserId": approval.UserId,
                    "ReviewerId": approval.ReviewerId,
                    "ApprovedNot": approval.ApprovedNot,
                    "main_policy_rejected": main_policy_rejected,
                    "rejected_subpolicies": rejected_subpolicies,
                    "is_compliance": False
                })
    return Response(result)
 
@api_view(['PUT'])
@permission_classes([AllowAny])
def submit_policy_review(request, approval_id):
    try:
        # Get the original approval
        approval = PolicyApproval.objects.get(ApprovalId=approval_id)
       
        # Validate and prepare data
        extracted_data = request.data.get('ExtractedData')
        if not extracted_data:
            return Response({'error': 'ExtractedData is required'}, status=status.HTTP_400_BAD_REQUEST)
       
        approved_not = request.data.get('ApprovedNot')
       
        # Simply create a new PolicyApproval object
        # Avoid using filters that might generate BINARY expressions
        new_version = "r1"  # Default version for reviewer
       
        # Try to determine the next version number without SQL LIKE
        try:
            r_versions = []
            for pa in PolicyApproval.objects.filter(Identifier=approval.Identifier):
                if pa.Version and pa.Version.startswith('r') and pa.Version[1:].isdigit():
                    r_versions.append(int(pa.Version[1:]))
           
            if r_versions:
                new_version = f"r{max(r_versions) + 1}"
        except Exception as version_err:
            print(f"Error determining version (using default): {str(version_err)}")
       
        # Set approved date if policy is approved
        approved_date = None
        if approved_not == True or approved_not == 1:
            approved_date = datetime.date.today()
           
        # Create a new record using Django ORM
        new_approval = PolicyApproval(
            Identifier=approval.Identifier,
            ExtractedData=extracted_data,
            UserId=approval.UserId,
            ReviewerId=approval.ReviewerId,
            ApprovedNot=approved_not,
            ApprovedDate=approved_date,  # Set approved date
            Version=new_version
        )
        new_approval.save()
       
        # If policy is approved (ApprovedNot=1), update the status in policy and subpolicies tables
        if approved_not == True or approved_not == 1:
            try:
                # Find the policy by Identifier
                policy = Policy.objects.get(Identifier=approval.Identifier)

                # Get the policy version record
                policy_version = PolicyVersion.objects.filter(
                    PolicyId=policy,
                    Version=policy.CurrentVersion
                ).first()

                # If this policy has a previous version, set it to inactive
                if policy_version and policy_version.PreviousVersionId:
                    try:
                        previous_version = PolicyVersion.objects.get(VersionId=policy_version.PreviousVersionId)
                        previous_policy = previous_version.PolicyId
                        previous_policy.ActiveInactive = 'Inactive'
                        previous_policy.save()
                        print(f"Set previous policy version {previous_policy.PolicyId} to Inactive")
                    except Exception as prev_error:
                        print(f"Error updating previous policy version: {str(prev_error)}")
               
                # Update policy status to Approved and Active
                if policy.Status == 'Under Review':
                    policy.Status = 'Approved'
                    policy.ActiveInactive = 'Active'  # Set to Active when approved
                    policy.save()
                    print(f"Updated policy {policy.Identifier} status to Approved and Active")
               
                # Update all subpolicies for this policy
                subpolicies = SubPolicy.objects.filter(PolicyId=policy.PolicyId)
                for subpolicy in subpolicies:
                    if subpolicy.Status == 'Under Review':
                        subpolicy.Status = 'Approved'
                        subpolicy.save()
                        print(f"Updated subpolicy {subpolicy.Identifier} status to Approved")
            except Exception as update_error:
                print(f"Error updating policy/subpolicy status: {str(update_error)}")
                # Continue with the response even if status update fails
       
        return Response({
            'message': 'Policy review submitted successfully',
            'ApprovalId': new_approval.ApprovalId,
            'Version': new_approval.Version,
            'ApprovedDate': approved_date.isoformat() if approved_date else None
        })
       
    except PolicyApproval.DoesNotExist:
        return Response({'error': 'Policy approval not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        print("Error in submit_policy_review:", str(e))
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)
    
@api_view(['POST'])
@permission_classes([AllowAny])
def add_subpolicy_to_policy(request, policy_id):
    policy = get_object_or_404(Policy, PolicyId=policy_id)
   
    try:
        with transaction.atomic():
            # Set policy ID and default values in the request data
            subpolicy_data = request.data.copy()
            subpolicy_data['PolicyId'] = policy.PolicyId
            if 'CreatedByName' not in subpolicy_data:
                subpolicy_data['CreatedByName'] = policy.CreatedByName
            if 'CreatedByDate' not in subpolicy_data:
                subpolicy_data['CreatedByDate'] = datetime.date.today()
            if 'Status' not in subpolicy_data:
                subpolicy_data['Status'] = 'Under Review'
           
            subpolicy_serializer = SubPolicySerializer(data=subpolicy_data)
            if not subpolicy_serializer.is_valid():
                return Response(subpolicy_serializer.errors, status=status.HTTP_400_BAD_REQUEST)
           
            subpolicy = subpolicy_serializer.save()
           
            return Response({
                'message': 'Subpolicy added to policy successfully',
                'SubPolicyId': subpolicy.SubPolicyId,
                'PolicyId': policy.PolicyId
            }, status=status.HTTP_201_CREATED)
    except Exception as e:
        error_info = {
            'error': str(e),
            'traceback': traceback.format_exc()
        }
        return Response({'error': 'Error adding subpolicy to policy', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)
 

"""
@api GET /api/subpolicies/{pk}/
Returns a specific subpolicy by ID if it has Status='Approved',
its parent policy has Status='Approved' and ActiveInactive='Active',
and its parent framework has Status='Approved' and ActiveInactive='Active'.

@api PUT /api/subpolicies/{pk}/
Updates an existing subpolicy. Only subpolicies with Status='Approved'
whose parent policy and framework are also Approved and Active can be updated.

Example payload:
{
  "SubPolicyName": "Enhanced Password Management",
  "Description": "Updated password requirements and management",
  "Control": "Use strong passwords with at least 16 characters, including special characters",
  "Identifier": "PWD-002",
}

@api DELETE /api/subpolicies/{pk}/
Soft-deletes a subpolicy by setting Status='Inactive'.
"""
@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([AllowAny])
def subpolicy_detail(request, pk):
    subpolicy = get_object_or_404(SubPolicy, SubPolicyId=pk)
    
    if request.method == 'GET':
        # Only return details if subpolicy is Approved
        if subpolicy.Status != 'Approved':
            return Response({'error': 'Subpolicy is not approved'}, status=status.HTTP_403_FORBIDDEN)
        
        # Get policy to check if it's approved and active
        policy = subpolicy.PolicyId
        if policy.Status != 'Approved' or policy.ActiveInactive != 'Active':
            return Response({'error': 'Policy is not approved or active'}, status=status.HTTP_403_FORBIDDEN)
        
        # Get framework to check if it's approved and active
        framework = policy.FrameworkId
        if framework.Status != 'Approved' or framework.ActiveInactive != 'Active':
            return Response({'error': 'Framework is not approved or active'}, status=status.HTTP_403_FORBIDDEN)
        
        serializer = SubPolicySerializer(subpolicy)
        return Response(serializer.data)
    
    elif request.method == 'PUT':
        # Check if subpolicy is approved before allowing update
        if subpolicy.Status != 'Approved':
            return Response({'error': 'Only approved subpolicies can be updated'}, status=status.HTTP_403_FORBIDDEN)
        
        # Check if policy is approved and active
        policy = subpolicy.PolicyId
        if policy.Status != 'Approved' or policy.ActiveInactive != 'Active':
            return Response({'error': 'Policy is not approved or active'}, status=status.HTTP_403_FORBIDDEN)
        
        # Check if framework is approved and active
        framework = policy.FrameworkId
        if framework.Status != 'Approved' or framework.ActiveInactive != 'Active':
            return Response({'error': 'Framework is not approved or active'}, status=status.HTTP_403_FORBIDDEN)
        
        try:
            with transaction.atomic():
                serializer = SubPolicySerializer(subpolicy, data=request.data, partial=True)
                if serializer.is_valid():
                    serializer.save()
                    return Response({
                        'message': 'Subpolicy updated successfully',
                        'SubPolicyId': subpolicy.SubPolicyId
                    })
                else:
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            error_info = {
                'error': str(e),
                'traceback': traceback.format_exc()
            }
            return Response({'error': 'Error updating subpolicy', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'DELETE':
        try:
            with transaction.atomic():
                # Instead of deleting, set Status to 'Inactive'
                subpolicy.Status = 'Inactive'
                subpolicy.save()
                
                return Response({'message': 'Subpolicy marked as inactive'}, status=status.HTTP_200_OK)
        except Exception as e:
            error_info = {
                'error': str(e),
                'traceback': traceback.format_exc()
            }
            return Response({'error': 'Error marking subpolicy as inactive', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)

"""
@api POST /api/frameworks/{pk}/copy/
Copies an existing framework to create a new one with modified details.
The FrameworkName must be unique - the request will be rejected if a framework with the same name already exists.
The copied framework will have Status='Under Review' and ActiveInactive='Inactive' by default.
All policies and subpolicies will be copied with the same structure but will also be set as Under Review/Inactive.
You can also modify specific policies by including a 'policies' array with PolicyId and updated fields.

Example payload:
{
  "FrameworkName": "ISO 27001:2023",
  "FrameworkDescription": "Updated Information Security Management System 2023 version",
  "EffectiveDate": "2023-11-01",
  "CreatedByName": "Jane Smith",
  "CreatedByDate": "2023-10-15",
  "Category": "Information Security and Compliance",
  "Identifier": "ISO-27001-2023",
  "policies": [
    {
      "original_policy_id": 1,
      "PolicyName": "Updated Access Control Policy 2023",
      "PolicyDescription": "Enhanced guidelines for access control with zero trust approach",
      "Department": "IT,Security",
      "Scope": "All IT systems and cloud environments",
      "Objective": "Implement zero trust security model"
    },
    {
      "original_policy_id": 2,
      "PolicyName": "Data Protection Policy 2023",
      "exclude": true
    }
  ]
}
"""
@api_view(['POST'])
@permission_classes([AllowAny])
def copy_framework(request, pk):
    # Get original framework
    original_framework = get_object_or_404(Framework, FrameworkId=pk)
    print(f"Original Framework: ID={original_framework.FrameworkId}, Name={original_framework.FrameworkName}")

    try:
        with transaction.atomic():
            # Verify original framework status
            print(f"Original Framework Status: {original_framework.Status}, ActiveInactive: {original_framework.ActiveInactive}")
            if original_framework.Status != 'Approved' or original_framework.ActiveInactive != 'Active':
                print("Original framework not Approved or Active - aborting copy.")
                return Response({
                    'error': 'Only Approved and Active frameworks can be copied'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Check framework name in request
            framework_name = request.data.get('FrameworkName')
            print(f"Requested new framework name: {framework_name}")
            if not framework_name:
                print("FrameworkName missing in request")
                return Response({'error': 'FrameworkName is required'}, status=status.HTTP_400_BAD_REQUEST)

            if Framework.objects.filter(FrameworkName=framework_name).exists():
                print("Framework with given name already exists.")
                return Response({'error': 'A framework with this name already exists'}, status=status.HTTP_400_BAD_REQUEST)

            framework_version = 1.0
            print(f"New framework version set to: {framework_version}")

            # Prepare new framework data
            new_framework_data = {
                'FrameworkName': framework_name,
                'CurrentVersion': framework_version,
                'FrameworkDescription': request.data.get('FrameworkDescription', original_framework.FrameworkDescription),
                'EffectiveDate': request.data.get('EffectiveDate', original_framework.EffectiveDate),
                'CreatedByName': request.data.get('CreatedByName', original_framework.CreatedByName),
                'CreatedByDate': datetime.date.today(),
                'Category': request.data.get('Category', original_framework.Category),
                'DocURL': request.data.get('DocURL', original_framework.DocURL),
                'Identifier': original_framework.Identifier,
                'StartDate': request.data.get('StartDate', original_framework.StartDate),
                'EndDate': request.data.get('EndDate', original_framework.EndDate),
                'Status': 'Under Review',
                'ActiveInactive': 'Inactive',
                'Reviewer': request.data.get('Reviewer', original_framework.Reviewer)
            }

            print(f"Creating new framework with data: {new_framework_data}")
            new_framework = Framework.objects.create(**new_framework_data)
            print(f"New Framework created: ID={new_framework.FrameworkId}, Name={new_framework.FrameworkName}")

            # Create framework version record
            framework_version_record = FrameworkVersion(
                FrameworkId=new_framework,
                Version=str(framework_version),
                FrameworkName=new_framework.FrameworkName,
                CreatedBy=new_framework.CreatedByName,
                CreatedDate=datetime.date.today(),
                PreviousVersionId=None
            )
            framework_version_record.save()
            print(f"FrameworkVersion record created for Framework ID {new_framework.FrameworkId} with Version {framework_version}")

            # Initialize policy tracking variables
            policy_customizations = {}
            policies_to_exclude = []
            created_policies = []

            # Handle policies from request
            if 'policies' in request.data:
                print(f"Received policies to process: {len(request.data.get('policies', []))}")
                for policy_data in request.data.get('policies', []):
                    if 'original_policy_id' in policy_data:
                        policy_id = policy_data.get('original_policy_id')
                        if policy_data.get('exclude', False):
                            print(f"Policy ID {policy_id} marked for exclusion")
                            policies_to_exclude.append(policy_id)
                        else:
                            print(f"Policy ID {policy_id} customization received")
                            policy_customizations[policy_id] = policy_data

            # Query original policies to copy
            original_policies = Policy.objects.filter(
                FrameworkId=original_framework,
                Status='Approved',
                ActiveInactive='Active'
            )
            print(f"Original policies count: {original_policies.count()}")

            for original_policy in original_policies:
                if original_policy.PolicyId in policies_to_exclude:
                    print(f"Skipping excluded policy: {original_policy.PolicyName} (ID {original_policy.PolicyId})")
                    continue
                print(f"Including policy: {original_policy.PolicyName} (ID {original_policy.PolicyId})")

                custom_data = policy_customizations.get(original_policy.PolicyId, {})

                created_by_user_id = custom_data.get('CreatedByUserId')
                if created_by_user_id:
                    try:
                        created_by_user = Users.objects.get(UserId=created_by_user_id)
                        created_by_name = created_by_user.UserName
                    except Users.DoesNotExist:
                        error_msg = f'User not found for CreatedByUserId: {created_by_user_id}'
                        print(error_msg)
                        return Response({'error': error_msg}, status=status.HTTP_400_BAD_REQUEST)
                else:
                    created_by_name = original_policy.CreatedByName

                reviewer_name = custom_data.get('Reviewer') or original_policy.Reviewer

                new_policy_data = {
                    'FrameworkId': new_framework,
                    'CurrentVersion': framework_version,
                    'Status': 'Under Review',
                    'PolicyDescription': custom_data.get('PolicyDescription', original_policy.PolicyDescription),
                    'PolicyName': custom_data.get('PolicyName', original_policy.PolicyName),
                    'StartDate': custom_data.get('StartDate', original_policy.StartDate),
                    'EndDate': custom_data.get('EndDate', original_policy.EndDate),
                    'Department': custom_data.get('Department', original_policy.Department),
                    'CreatedByName': custom_data.get('CreatedByName', created_by_name),
                    'CreatedByDate': new_framework.CreatedByDate,
                    'Applicability': custom_data.get('Applicability', original_policy.Applicability),
                    'DocURL': custom_data.get('DocURL', original_policy.DocURL),
                    'Scope': custom_data.get('Scope', original_policy.Scope),
                    'Objective': custom_data.get('Objective', original_policy.Objective),
                    'Identifier': custom_data.get('Identifier', original_policy.Identifier),
                    'PermanentTemporary': custom_data.get('PermanentTemporary', original_policy.PermanentTemporary),
                    'ActiveInactive': 'Inactive',
                    'Reviewer': reviewer_name,
                    'CoverageRate': custom_data.get('CoverageRate', original_policy.CoverageRate)
                }

                print(f"Creating new policy with data: {new_policy_data}")
                new_policy = Policy.objects.create(**new_policy_data)
                created_policies.append(new_policy)
                print(f"Created policy: {new_policy.PolicyName} (ID {new_policy.PolicyId})")

                # Subpolicy handling initialization
                subpolicy_customizations = {}
                subpolicies_to_exclude = []

                if 'subpolicies' in custom_data:
                    print(f"Policy {new_policy.PolicyName} has subpolicies to process: {len(custom_data.get('subpolicies', []))}")
                    for subpolicy_data in custom_data.get('subpolicies', []):
                        if 'original_subpolicy_id' in subpolicy_data:
                            subpolicy_id = subpolicy_data.get('original_subpolicy_id')
                            if subpolicy_data.get('exclude', False):
                                print(f"Subpolicy ID {subpolicy_id} marked for exclusion")
                                subpolicies_to_exclude.append(subpolicy_id)
                            else:
                                print(f"Customization for subpolicy ID {subpolicy_id} received")
                                subpolicy_customizations[subpolicy_id] = subpolicy_data

                original_subpolicies = SubPolicy.objects.filter(PolicyId=original_policy)
                print(f"Original subpolicies count for policy {original_policy.PolicyName}: {original_subpolicies.count()}")

                for subpolicy in original_subpolicies:
                    if subpolicy.SubPolicyId not in subpolicies_to_exclude:
                        sub_custom_data = subpolicy_customizations.get(subpolicy.SubPolicyId, {})
                        new_subpolicy_data = {
                            'PolicyId': new_policy,
                            'SubPolicyName': sub_custom_data.get('SubPolicyName', subpolicy.SubPolicyName),
                            'CreatedByName': new_policy.CreatedByName,
                            'CreatedByDate': datetime.date.today(),
                            'Identifier': sub_custom_data.get('Identifier', subpolicy.Identifier),
                            'Description': sub_custom_data.get('Description', subpolicy.Description),
                            'Status': 'Under Review',
                            'PermanentTemporary': sub_custom_data.get('PermanentTemporary', subpolicy.PermanentTemporary),
                            'Control': sub_custom_data.get('Control', subpolicy.Control)
                        }
                        SubPolicy.objects.create(**new_subpolicy_data)
                        print(f"Created subpolicy: {new_subpolicy_data['SubPolicyName']} for policy {new_policy.PolicyName}")

            if 'new_policies' in request.data:
                print(f"Processing {len(request.data.get('new_policies', []))} new policies")
                for new_policy_data in request.data.get('new_policies', []):
                    required_fields = ['PolicyName', 'PolicyDescription', 'Identifier']
                    missing_fields = [f for f in required_fields if f not in new_policy_data]
                    if missing_fields:
                        error_msg = f"Missing required fields for new policy: {', '.join(missing_fields)}"
                        print(error_msg)
                        return Response({'error': error_msg}, status=status.HTTP_400_BAD_REQUEST)

                    subpolicies_data = new_policy_data.pop('subpolicies', [])

                    policy_data = new_policy_data.copy()
                    policy_data['FrameworkId'] = new_framework
                    policy_data['CurrentVersion'] = framework_version
                    policy_data['Status'] = 'Under Review'
                    policy_data['ActiveInactive'] = 'Inactive'
                    policy_data.setdefault('CreatedByName', new_framework.CreatedByName)
                    policy_data['CreatedByDate'] = datetime.date.today()

                    new_policy = Policy.objects.create(**policy_data)
                    created_policies.append(new_policy)
                    print(f"Created new policy: {new_policy.PolicyName} (ID {new_policy.PolicyId})")

                    PolicyVersion.objects.create(
                        PolicyId=new_policy,
                        Version=str(framework_version),
                        PolicyName=new_policy.PolicyName,
                        CreatedBy=new_policy.CreatedByName,
                        CreatedDate=datetime.date.today(),
                        PreviousVersionId=None
                    )

                    for subpolicy_data in subpolicies_data:
                        required_fields = ['SubPolicyName', 'Description', 'Identifier']
                        missing_fields = [f for f in required_fields if f not in subpolicy_data]
                        if missing_fields:
                            error_msg = f"Missing required fields for subpolicy in new policy {new_policy.PolicyName}: {', '.join(missing_fields)}"
                            print(error_msg)
                            return Response({'error': error_msg}, status=status.HTTP_400_BAD_REQUEST)

                        subpolicy = subpolicy_data.copy()
                        subpolicy['PolicyId'] = new_policy
                        subpolicy.setdefault('CreatedByName', new_policy.CreatedByName)
                        subpolicy['CreatedByDate'] = datetime.date.today()
                        subpolicy.setdefault('Status', 'Under Review')

                        SubPolicy.objects.create(**subpolicy)
                        print(f"Created subpolicy: {subpolicy.get('SubPolicyName')} for new policy {new_policy.PolicyName}")

            response_data = {
                'message': 'Framework copied successfully',
                'FrameworkId': new_framework.FrameworkId,
                'FrameworkName': new_framework.FrameworkName,
                'Version': new_framework.CurrentVersion,
            }

            if created_policies:
                response_data['policies'] = [{
                    'PolicyId': p.PolicyId,
                    'PolicyName': p.PolicyName,
                    'Identifier': p.Identifier,
                    'Version': p.CurrentVersion
                } for p in created_policies]

            print(f"Copy framework operation completed successfully for Framework ID {new_framework.FrameworkId}")
            return Response(response_data, status=status.HTTP_201_CREATED)

    except Exception as e:
        error_info = {
            'error': str(e),
            'traceback': traceback.format_exc()
        }
        print("Error in copy_framework:", error_info)
        return Response({'error': 'Error copying framework', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)


"""
@api POST /api/policies/{pk}/copy/
Copies an existing policy to create a new one with modified details within the same framework.
The PolicyName must be unique within the framework - the request will be rejected if a policy with the same name already exists.
The copied policy will have Status='Under Review' and ActiveInactive='Inactive' by default.
All subpolicies will be copied with the same structure but will also be set as Under Review by default.
You can also modify, exclude, or add new subpolicies.

Example payload:
{
  "PolicyName": "Enhanced Access Control Policy 2023",
  "PolicyDescription": "Updated guidelines for access control with zero trust approach",
  "StartDate": "2023-11-01",
  "EndDate": "2025-11-01",
  "Department": "IT,Security",
  "CreatedByName": "Jane Smith",
  "CreatedByDate": "2023-10-15",
  "Scope": "All IT systems and cloud environments",
  "Objective": "Implement zero trust security model",
  "Identifier": "ACP-ZT-001",
  "subpolicies": [
    {
      "original_subpolicy_id": 5,
      "SubPolicyName": "Enhanced Password Rules",
      "Description": "Updated password requirements with MFA",
      "Control": "16-character passwords with MFA for all access"
    },
    {
      "original_subpolicy_id": 6,
      "exclude": true
    }
  ],
  "new_subpolicies": [
    {
      "SubPolicyName": "Device Authentication",
      "Description": "Requirements for device-based authentication",
      "Control": "Implement device certificates for all company devices",
      "Identifier": "DEV-AUTH-001",
      "Status": "Under Review"
    }
  ]
}
"""
@api_view(['POST'])
@permission_classes([AllowAny])
def copy_policy(request, pk):
    # Get original policy
    original_policy = get_object_or_404(Policy, PolicyId=pk)
    
    try:
        with transaction.atomic():
            # Verify the original policy is Approved and Active
            if original_policy.Status != 'Approved' or original_policy.ActiveInactive != 'Active':
                return Response({
                    'error': 'Only Approved and Active policies can be copied'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Check if policy name is unique within the framework
            policy_name = request.data.get('PolicyName')
            if not policy_name:
                return Response({'error': 'PolicyName is required'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Get target framework ID from request
            target_framework_id = request.data.get('TargetFrameworkId')
            if not target_framework_id:
                return Response({'error': 'TargetFrameworkId is required'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Get target framework
            try:
                target_framework = Framework.objects.get(FrameworkId=target_framework_id)
            except Framework.DoesNotExist:
                return Response({'error': 'Target framework not found'}, status=status.HTTP_404_NOT_FOUND)
            
            # Check if a policy with this name already exists in the target framework
            if Policy.objects.filter(FrameworkId=target_framework, PolicyName=policy_name).exists():
                return Response({'error': 'A policy with this name already exists in the target framework'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Create new policy with data from original and overrides from request
            new_policy_data = {
                'FrameworkId': target_framework,  # Use target framework instead of original
                'Status': 'Under Review',
                'PolicyName': policy_name,
                'PolicyDescription': request.data.get('PolicyDescription', original_policy.PolicyDescription),
                'StartDate': request.data.get('StartDate', original_policy.StartDate),
                'EndDate': request.data.get('EndDate', original_policy.EndDate),
                'Department': request.data.get('Department', original_policy.Department),
                'CreatedByName': request.data.get('CreatedByName', original_policy.CreatedByName),
                'CreatedByDate': request.data.get('CreatedByDate', datetime.date.today()),
                'Applicability': request.data.get('Applicability', original_policy.Applicability),
                'DocURL': request.data.get('DocURL', original_policy.DocURL),
                'Scope': request.data.get('Scope', original_policy.Scope),
                'Objective': request.data.get('Objective', original_policy.Objective),
                'Identifier': request.data.get('Identifier', original_policy.Identifier),
                'PermanentTemporary': request.data.get('PermanentTemporary', original_policy.PermanentTemporary),
                'ActiveInactive': 'Inactive',
                'CurrentVersion': 1.0,  # Start with version 1.0 for new policy
                'Reviewer': request.data.get('Reviewer', original_policy.Reviewer),  # Store reviewer name
                'CoverageRate': request.data.get('CoverageRate', original_policy.CoverageRate)  # Add coverage rate
            }
            
            # Create new policy
            new_policy = Policy.objects.create(**new_policy_data)
            
            # Create policy version record (no previous version link) - ONLY ONCE
            policy_version = PolicyVersion(
                PolicyId=new_policy,
                Version='1.0',  # Start with version 1.0
                PolicyName=new_policy.PolicyName,
                CreatedBy=new_policy.CreatedByName,
                CreatedDate=new_policy.CreatedByDate,
                PreviousVersionId=None  # No version linking
            )
            policy_version.save()
            
            # Handle subpolicy customizations if provided
            subpolicy_customizations = {}
            subpolicies_to_exclude = []
            created_subpolicies = []  # Keep track of created subpolicies
            
            # Process subpolicy customizations if provided
            if 'subpolicies' in request.data:
                for subpolicy_data in request.data.get('subpolicies', []):
                    if 'original_subpolicy_id' in subpolicy_data:
                        subpolicy_id = subpolicy_data.get('original_subpolicy_id')
                        
                        # Check if this subpolicy should be excluded
                        if subpolicy_data.get('exclude', False):
                            subpolicies_to_exclude.append(subpolicy_id)
                        else:
                            # Store customizations for this subpolicy
                            subpolicy_customizations[subpolicy_id] = subpolicy_data
            
            # Copy only Approved and Active subpolicies from original policy - ONLY ONCE
            original_subpolicies = SubPolicy.objects.filter(
                PolicyId=original_policy,
                Status='Approved'
            )
            
            for original_subpolicy in original_subpolicies:
                # Skip if this subpolicy should be excluded
                if original_subpolicy.SubPolicyId in subpolicies_to_exclude:
                    continue
                
                # Get customizations for this subpolicy if any
                custom_data = subpolicy_customizations.get(original_subpolicy.SubPolicyId, {})
                
                # Create new subpolicy with data from original and any customizations
                new_subpolicy_data = {
                    'PolicyId': new_policy,
                    'SubPolicyName': custom_data.get('SubPolicyName', original_subpolicy.SubPolicyName),
                    'CreatedByName': new_policy.CreatedByName,
                    'CreatedByDate': new_policy.CreatedByDate,
                    'Identifier': custom_data.get('Identifier', original_subpolicy.Identifier),
                    'Description': custom_data.get('Description', original_subpolicy.Description),
                    'Status': 'Under Review',
                    'PermanentTemporary': custom_data.get('PermanentTemporary', original_subpolicy.PermanentTemporary),
                    'Control': custom_data.get('Control', original_subpolicy.Control)
                }
                
                new_subpolicy = SubPolicy.objects.create(**new_subpolicy_data)
                created_subpolicies.append(new_subpolicy)
            
            # --- Add this block to process new subpolicies (no original_subpolicy_id) ---
            if 'subpolicies' in request.data:
                for subpolicy_data in request.data.get('subpolicies', []):
                    if 'original_subpolicy_id' not in subpolicy_data:
                        # This is a new subpolicy, create it
                        new_subpolicy_data = {
                            'PolicyId': new_policy,
                            'SubPolicyName': subpolicy_data.get('SubPolicyName'),
                            'CreatedByName': new_policy.CreatedByName,
                            'CreatedByDate': new_policy.CreatedByDate,
                            'Identifier': subpolicy_data.get('Identifier'),
                            'Description': subpolicy_data.get('Description'),
                            'Status': 'Under Review',
                            'PermanentTemporary': subpolicy_data.get('PermanentTemporary', 'Permanent'),
                            'Control': subpolicy_data.get('Control')
                        }
                        SubPolicy.objects.create(**new_subpolicy_data)
            # --- End new subpolicy block ---
            
            return Response({
                'message': 'Policy copied successfully to target framework',
                'PolicyId': new_policy.PolicyId,
                'PolicyName': new_policy.PolicyName,
                'SourceFrameworkId': original_policy.FrameworkId.FrameworkId,
                'TargetFrameworkId': target_framework.FrameworkId,
                'Version': new_policy.CurrentVersion
            }, status=status.HTTP_201_CREATED)
    except Exception as e:
        error_info = {
            'error': str(e),
            'traceback': traceback.format_exc()
        }
        print("Error in copy_policy:", error_info)  # Add this to see full error on server console/logs
        return Response({'error': 'Error copying policy', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)

"""
@api PUT /api/frameworks/{pk}/toggle-status/
Toggles the ActiveInactive status of a framework between 'Active' and 'Inactive'.
If the framework is currently 'Active', it will be set to 'Inactive' and vice versa.
When a framework is set to 'Inactive', all its policies will also be set to 'Inactive'.

Example response:
{
    "message": "Framework status updated successfully",
    "FrameworkId": 1,
    "FrameworkName": "ISO 27001",
    "ActiveInactive": "Inactive"
}
"""
@api_view(['PUT'])
@permission_classes([AllowAny])
def toggle_framework_status(request, pk):
    framework = get_object_or_404(Framework, FrameworkId=pk)
    
    try:
        with transaction.atomic():
            # Toggle the status
            new_status = 'Inactive' if framework.ActiveInactive == 'Active' else 'Active'
            framework.ActiveInactive = new_status
            framework.save()
            
            # If setting to Inactive, also set all policies to Inactive
            if new_status == 'Inactive':
                policies = Policy.objects.filter(FrameworkId=framework)
                for policy in policies:
                    policy.ActiveInactive = 'Inactive'
                    policy.save()
            
            return Response({
                'message': 'Framework status updated successfully',
                'FrameworkId': framework.FrameworkId,
                'FrameworkName': framework.FrameworkName,
                'ActiveInactive': framework.ActiveInactive
            })
    except Exception as e:
        error_info = {
            'error': str(e),
            'traceback': traceback.format_exc()
        }
        return Response({'error': 'Error updating framework status', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)

"""
@api PUT /api/policies/{pk}/toggle-status/
Toggles the ActiveInactive status of a policy between 'Active' and 'Inactive'.
If the policy is currently 'Active', it will be set to 'Inactive' and vice versa.
Note: A policy can only be set to 'Active' if its parent framework is also 'Active'.

Example response:
{
    "message": "Policy status updated successfully",
    "PolicyId": 1,
    "PolicyName": "Access Control Policy",
    "ActiveInactive": "Active"
}
"""
@api_view(['PUT'])
@permission_classes([AllowAny])
def toggle_policy_status(request, pk):
    policy = get_object_or_404(Policy, PolicyId=pk)
    
    try:
        with transaction.atomic():
            # Check if trying to activate a policy while framework is inactive
            if policy.ActiveInactive == 'Inactive' and policy.FrameworkId.ActiveInactive == 'Inactive':
                return Response({
                    'error': 'Cannot activate policy while parent framework is inactive'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Toggle the status
            new_status = 'Inactive' if policy.ActiveInactive == 'Active' else 'Active'
            policy.ActiveInactive = new_status
            policy.save()
            
            return Response({
                'message': 'Policy status updated successfully',
                'PolicyId': policy.PolicyId,
                'PolicyName': policy.PolicyName,
                'ActiveInactive': policy.ActiveInactive
            })
    except Exception as e:
        error_info = {
            'error': str(e),
            'traceback': traceback.format_exc()
        }
        return Response({'error': 'Error updating policy status', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)

"""
@api POST /api/frameworks/{pk}/create-version/
Creates a new version of an existing framework by cloning it with an incremented version number.
For example, if the original framework has version 1.0, the new version will be 1.1.
All policies and subpolicies will be cloned with their details.

Example payload:
{
  "FrameworkName": "ISO 27001 v3.3",
  "FrameworkDescription": "Updated Information Security Management System 2024",
  "EffectiveDate": "2024-01-01",
  "CreatedByName": "John Doe",
  "CreatedByDate": "2023-12-15",
  "policies": [
    {
      "original_policy_id": 1052,
      "PolicyName": "Access Control Policy",
      "PolicyDescription": "Original access control policy",
      "Identifier": "ACP-001",
      "subpolicies": [
        {
          "original_subpolicy_id": 100,
          "SubPolicyName": "Password Management",
          "Description": "Original password requirements",
          "Control": "Use strong passwords",
          "Identifier": "PWD-001"
        }
      ]
    },
    {
      "original_policy_id": 2,
      "exclude": true
    },
    {
      "original_policy_id": 3,
      "PolicyName": "Data Protection Policy",
      "PolicyDescription": "Original data protection policy",
      "Identifier": "DPP-001",
      "subpolicies": [
        {
          "original_subpolicy_id": 4,
          "exclude": true
        },
        {
          "original_subpolicy_id": 5,
          "exclude": true
        }
      ]
    }
  ],
  "new_policies": [
    {
      "PolicyName": "New Security Policy",
      "PolicyDescription": "A completely new policy",
      "Identifier": "NSP-001",
      "Department": "IT,Security",
      "Scope": "All systems",
      "Objective": "Implement new security measures",
      "subpolicies": [
        {
          "SubPolicyName": "New Security Control",
          "Description": "New security requirements",
          "Control": "Implement new security measures",
          "Identifier": "NSC-001"
        }
      ]
    }
  ]
}

Example response:
{
    "message": "New framework version created successfully",
    "FrameworkId": 35,
    "FrameworkName": "ISO 27001 v3.3",
    "PreviousVersion": 1.0,
    "NewVersion": 1.1,
    "Identifier": "ISO",
    "policies": [
        {
            "PolicyId": 1074,
            "PolicyName": "Access Control Policy",
            "Identifier": "ACP-001",
            "Version": 1.1
        },
        {
            "PolicyId": 1075,
            "PolicyName": "New Security Policy",
            "Identifier": "NSP-001",
            "Version": 1.1
        }
    ]
}
"""
@api_view(['POST'])
@permission_classes([AllowAny])
def create_framework_version(request, pk):
    original_framework = get_object_or_404(Framework, FrameworkId=pk)
    
    try:
        with transaction.atomic():
            reviewer_id = request.data.get('Reviewer')
            reviewer_name = None
            if reviewer_id:
                user_obj = Users.objects.filter(UserId=reviewer_id).first()
                if user_obj:
                    reviewer_name = user_obj.UserName

            framework_name = request.data.get('FrameworkName')
            if not framework_name:
                return Response({'error': 'FrameworkName is required'}, status=status.HTTP_400_BAD_REQUEST)
            
            latest_version = FrameworkVersion.objects.filter(
                FrameworkId__Identifier=original_framework.Identifier
            ).order_by('-Version').first()
            
            current_version = int(float(original_framework.CurrentVersion)) if not latest_version else int(float(latest_version.Version))
            new_version = current_version + 1
            new_version_str = f"{new_version}.0"
            
            new_framework_data = {
                'FrameworkName': framework_name,
                'CurrentVersion': new_version_str,
                'FrameworkDescription': request.data.get('FrameworkDescription', original_framework.FrameworkDescription),
                'EffectiveDate': request.data.get('EffectiveDate', original_framework.EffectiveDate),
                'CreatedByName': request.data.get('CreatedByName', original_framework.CreatedByName),
                'CreatedByDate': request.data.get('CreatedByDate', datetime.date.today()),
                'Category': request.data.get('Category', original_framework.Category),
                'DocURL': request.data.get('DocURL', original_framework.DocURL),
                'Identifier': original_framework.Identifier,
                'StartDate': request.data.get('StartDate', original_framework.StartDate),
                'EndDate': request.data.get('EndDate', original_framework.EndDate),
                'Status': 'Under Review',
                'ActiveInactive': 'Inactive',
                'Reviewer': reviewer_name,
            }
            
            new_framework = Framework.objects.create(**new_framework_data)

            original_framework_version = FrameworkVersion.objects.filter(
                FrameworkId=original_framework,
                Version=str(original_framework.CurrentVersion)
            ).first()
            
            if not original_framework_version:
                original_framework_version = FrameworkVersion.objects.create(
                    FrameworkId=original_framework,
                    Version=str(original_framework.CurrentVersion),
                    FrameworkName=original_framework.FrameworkName,
                    CreatedBy=original_framework.CreatedByName,
                    CreatedDate=original_framework.CreatedByDate,
                    PreviousVersionId=None
                )

            framework_version = FrameworkVersion(
                FrameworkId=new_framework,
                Version=str(new_version_str),
                FrameworkName=new_framework.FrameworkName,
                CreatedBy=new_framework.CreatedByName,
                CreatedDate=new_framework.CreatedByDate,
                PreviousVersionId=original_framework_version.VersionId
            )
            framework_version.save()

            # Process policies
            policy_customizations = {}
            policies_to_exclude = []
            created_policies = []

            if 'policies' in request.data:
                for policy_data in request.data.get('policies', []):
                    if 'original_policy_id' in policy_data:
                        policy_id = policy_data.get('original_policy_id')
                        if policy_data.get('exclude', False):
                            policies_to_exclude.append(int(policy_id))
                        else:
                            policy_customizations[int(policy_id)] = policy_data
            
            print("Policies to exclude:", policies_to_exclude)

            original_policies = Policy.objects.filter(
                FrameworkId=original_framework,
                Status='Approved',
                ActiveInactive='Active'
            )

            excluded_set = set(policies_to_exclude)

            for original_policy in original_policies:
                if original_policy.PolicyId in excluded_set:
                    print(f"Skipping excluded policy: {original_policy.PolicyName} (ID {original_policy.PolicyId})")
                    continue
                print(f"Including policy: {original_policy.PolicyName} (ID {original_policy.PolicyId})")

                custom_data = policy_customizations.get(original_policy.PolicyId, {})

                # Determine Reviewer username for existing policy
                reviewer_id_from_custom = custom_data.get('Reviewer')
                reviewer_username = reviewer_name  # fallback
                
                if reviewer_id_from_custom:
                    user_obj = Users.objects.filter(UserId=reviewer_id_from_custom).first()
                    if user_obj:
                        reviewer_username = user_obj.UserName

                new_policy_data = {
                    'FrameworkId': new_framework,
                    'CurrentVersion': new_version_str,
                    'Status': 'Under Review',
                    'PolicyDescription': custom_data.get('PolicyDescription', original_policy.PolicyDescription),
                    'PolicyName': custom_data.get('PolicyName', original_policy.PolicyName),
                    'StartDate': custom_data.get('StartDate', original_policy.StartDate),
                    'EndDate': custom_data.get('EndDate', original_policy.EndDate),
                    'Department': custom_data.get('Department', original_policy.Department),
                    'CreatedByName': custom_data.get('CreatedByName', original_policy.CreatedByName),
                    'CreatedByDate': new_framework.CreatedByDate,
                    'Applicability': custom_data.get('Applicability', original_policy.Applicability),
                    'DocURL': custom_data.get('DocURL', original_policy.DocURL),
                    'Scope': custom_data.get('Scope', original_policy.Scope),
                    'Objective': custom_data.get('Objective', original_policy.Objective),
                    'Identifier': custom_data.get('Identifier', original_policy.Identifier),
                    'PermanentTemporary': custom_data.get('PermanentTemporary', original_policy.PermanentTemporary),
                    'ActiveInactive': 'Inactive',
                    'Reviewer': reviewer_username,
                }
                
                new_policy = Policy.objects.create(**new_policy_data)
                created_policies.append(new_policy)

                original_policy_version = PolicyVersion.objects.filter(
                    PolicyId=original_policy,
                    Version=str(original_policy.CurrentVersion)
                ).first()

                if not original_policy_version:
                    return Response({
                        'error': f'No PolicyVersion found for PolicyId={original_policy.PolicyId} and Version={original_policy.CurrentVersion}. Data integrity issue.'
                    }, status=status.HTTP_400_BAD_REQUEST)

                policy_version = PolicyVersion(
                    PolicyId=new_policy,
                    Version=str(new_version_str),
                    PolicyName=new_policy.PolicyName,
                    CreatedBy=new_policy.CreatedByName,
                    CreatedDate=new_policy.CreatedByDate,
                    PreviousVersionId=original_policy_version.VersionId
                )
                policy_version.save()

                # Handle subpolicies
                subpolicy_customizations = {}
                subpolicies_to_exclude = []

                if 'subpolicies' in custom_data:
                    for subpolicy_data in custom_data.get('subpolicies', []):
                        if subpolicy_data.get('exclude', False):
                            if 'original_subpolicy_id' in subpolicy_data:
                                subpolicies_to_exclude.append(int(subpolicy_data.get('original_subpolicy_id')))
                        else:
                            if 'original_subpolicy_id' in subpolicy_data:
                                subpolicy_customizations[int(subpolicy_data.get('original_subpolicy_id'))] = subpolicy_data
                
                print(f"Subpolicies to exclude for policy {new_policy.PolicyName}:", subpolicies_to_exclude)

                original_subpolicies = SubPolicy.objects.filter(PolicyId=original_policy)

                excluded_subpolicy_set = set(subpolicies_to_exclude)
                for original_subpolicy in original_subpolicies:
                    if original_subpolicy.SubPolicyId in excluded_subpolicy_set:
                        print(f"Skipping excluded subpolicy: {original_subpolicy.SubPolicyName} (ID {original_subpolicy.SubPolicyId})")
                        continue
                    
                    custom_subpolicy_data = subpolicy_customizations.get(original_subpolicy.SubPolicyId, {})

                    new_subpolicy_data = {
                        'PolicyId': new_policy,
                        'SubPolicyName': custom_subpolicy_data.get('SubPolicyName', original_subpolicy.SubPolicyName),
                        'CreatedByName': new_policy.CreatedByName,
                        'CreatedByDate': new_policy.CreatedByDate,
                        'Identifier': custom_subpolicy_data.get('Identifier', original_subpolicy.Identifier),
                        'Description': custom_subpolicy_data.get('Description', original_subpolicy.Description),
                        'Status': 'Under Review',
                        'PermanentTemporary': custom_subpolicy_data.get('PermanentTemporary', original_subpolicy.PermanentTemporary),
                        'Control': custom_subpolicy_data.get('Control', original_subpolicy.Control)
                    }

                    SubPolicy.objects.create(**new_subpolicy_data)

                for sp_data in custom_data.get('subpolicies', []):
                    if sp_data.get('exclude', False):
                        continue
                    if 'original_subpolicy_id' in sp_data:
                        continue

                    required_fields = ['SubPolicyName', 'Description', 'Identifier']
                    missing_fields = [field for field in required_fields if field not in sp_data]
                    if missing_fields:
                        return Response({
                            'error': f'Missing required fields for new subpolicy in existing policy {new_policy.PolicyName}: {", ".join(missing_fields)}'
                        }, status=status.HTTP_400_BAD_REQUEST)

                    new_subpolicy_data = {
                        'PolicyId': new_policy,
                        'SubPolicyName': sp_data.get('SubPolicyName'),
                        'Description': sp_data.get('Description'),
                        'Identifier': sp_data.get('Identifier'),
                        'CreatedByName': new_policy.CreatedByName,
                        'CreatedByDate': new_policy.CreatedByDate,
                        'Status': 'Under Review',
                        'PermanentTemporary': sp_data.get('PermanentTemporary', 'Permanent'),
                        'Control': sp_data.get('Control', '')
                    }

                    SubPolicy.objects.create(**new_subpolicy_data)

            # Handle new policies
            if 'new_policies' in request.data:
                for new_policy_data in request.data.get('new_policies', []):
                    required_fields = ['PolicyName', 'PolicyDescription', 'Identifier']
                    missing_fields = [field for field in required_fields if field not in new_policy_data]
                    
                    if missing_fields:
                        return Response({
                            'error': f'Missing required fields for new policy: {", ".join(missing_fields)}'
                        }, status=status.HTTP_400_BAD_REQUEST)
                    
                    subpolicies_data = new_policy_data.pop('subpolicies', [])
                    
                    policy_data = new_policy_data.copy()
                    policy_data['FrameworkId'] = new_framework
                    policy_data['CurrentVersion'] = new_version_str
                    policy_data['Status'] = 'Under Review'
                    policy_data['ActiveInactive'] = 'Inactive'
                    policy_data.setdefault('CreatedByName', new_framework.CreatedByName)
                    policy_data['CreatedByDate'] = datetime.date.today()

                    reviewer_id_new_policy = policy_data.get('Reviewer')
                    reviewer_username_new_policy = reviewer_name  # fallback
                    if reviewer_id_new_policy:
                        user_obj = Users.objects.filter(UserId=reviewer_id_new_policy).first()
                        if user_obj:
                            reviewer_username_new_policy = user_obj.UserName

                    policy_data['Reviewer'] = reviewer_username_new_policy
                    
                    created_policy = Policy.objects.create(**policy_data)
                    created_policies.append(created_policy)
                    
                    PolicyVersion.objects.create(
                        PolicyId=created_policy,
                        Version=new_version_str,
                        PolicyName=created_policy.PolicyName,
                        CreatedBy=created_policy.CreatedByName,
                        CreatedDate=created_policy.CreatedByDate,
                        PreviousVersionId=None
                    )
                    
                    for subpolicy_data in subpolicies_data:
                        if subpolicy_data.get('exclude', False):
                            continue
                        
                        required_fields = ['SubPolicyName', 'Description', 'Identifier']
                        missing_fields = [field for field in required_fields if field not in subpolicy_data]
                        
                        if missing_fields:
                            return Response({
                                'error': f'Missing required fields for subpolicy in new policy {created_policy.PolicyName}: {", ".join(missing_fields)}'
                            }, status=status.HTTP_400_BAD_REQUEST)
                        
                        subpolicy = subpolicy_data.copy()
                        subpolicy['PolicyId'] = created_policy
                        subpolicy.setdefault('CreatedByName', created_policy.CreatedByName)
                        subpolicy.setdefault('CreatedByDate', created_policy.CreatedByDate)
                        subpolicy.setdefault('Status', 'Under Review')
                        
                        SubPolicy.objects.create(**subpolicy)
            
            print(f"Created policies count: {len(created_policies)}")
            
            response_data = {
                'message': 'New framework version created successfully',
                'FrameworkId': new_framework.FrameworkId,
                'FrameworkName': new_framework.FrameworkName,
                'PreviousVersion': current_version,
                'NewVersion': new_version,
                'Identifier': new_framework.Identifier,
            }
            
            if created_policies:
                response_data['policies'] = [{
                    'PolicyId': policy.PolicyId,
                    'PolicyName': policy.PolicyName,
                    'Identifier': policy.Identifier,
                    'Version': policy.CurrentVersion
                } for policy in created_policies]
            
            return Response(response_data, status=status.HTTP_201_CREATED)

    except Exception as e:
        error_info = {
            'error': str(e),
            'traceback': traceback.format_exc()
        }
        return Response({'error': 'Error creating new framework version', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)


"""
@api POST /api/policies/{pk}/create-version/
Creates a new version of an existing policy by cloning it with an incremented version number.
For example, if the original policy has version 1.0, the new version will be 1.1.
All subpolicies will be cloned with their details.
 
Example payload:
{
  "PolicyName": "Access Control Policy v1.1",
  "PolicyDescription": "Updated guidelines for access control",
  "StartDate": "2024-01-01",
  "EndDate": "2025-01-01",
  "Department": "IT,Security",
  "CreatedByName": "John Doe",
  "CreatedByDate": "2023-12-15",
  "Scope": "All IT systems and cloud environments",
  "Objective": "Implement enhanced access control measures"
}
 
Example response:
{
    "message": "New policy version created successfully",
    "PolicyId": 2,
    "PolicyName": "Access Control Policy v1.1",
    "PreviousVersion": 1.0,
    "NewVersion": 1.1,
    "FrameworkId": 1
}
"""
@api_view(['POST'])
@permission_classes([AllowAny])
def create_policy_version(request, pk):
    original_policy = get_object_or_404(Policy, PolicyId=pk)

    try:
        with transaction.atomic():
            # Validate policy name
            policy_name = request.data.get('PolicyName')
            if not policy_name:
                return Response({'error': 'PolicyName is required'}, status=status.HTTP_400_BAD_REQUEST)

            current_version = str(original_policy.CurrentVersion).strip()

            # Find latest minor version under same major version
            latest_version = PolicyVersion.objects.filter(
                PolicyId__Identifier=original_policy.Identifier,
                Version__startswith=current_version.split('.')[0] + '.'
            ).order_by('-Version').first()

            if latest_version:
                parts = latest_version.Version.split('.')
                if len(parts) == 2:
                    major, minor = parts[0], int(parts[1])
                    new_version = f"{major}.{minor + 1}"
                else:
                    new_version = f"{current_version}.1"
            else:
                new_version = f"{current_version.split('.')[0]}.1"

            # Resolve Reviewer UserName from UserId if given in request, fallback to original
            reviewer_id = request.data.get('Reviewer')
            reviewer_name = None
            if reviewer_id:
                user_obj = Users.objects.filter(UserId=reviewer_id).first()
                if user_obj:
                    reviewer_name = user_obj.UserName
            if not reviewer_name:
                reviewer_name = original_policy.Reviewer  # fallback to existing username

            # Prepare new policy data with Reviewer as UserName
            new_policy_data = {
                'FrameworkId': original_policy.FrameworkId,
                'CurrentVersion': new_version,
                'Status': 'Under Review',
                'PolicyName': policy_name,
                'PolicyDescription': request.data.get('PolicyDescription', original_policy.PolicyDescription),
                'StartDate': request.data.get('StartDate', original_policy.StartDate),
                'EndDate': request.data.get('EndDate', original_policy.EndDate),
                'Department': request.data.get('Department', original_policy.Department),
                'CreatedByName': request.data.get('CreatedByName', original_policy.CreatedByName),
                'CreatedByDate': datetime.date.today(),
                'Applicability': request.data.get('Applicability', original_policy.Applicability),
                'DocURL': request.data.get('DocURL', original_policy.DocURL),
                'Scope': request.data.get('Scope', original_policy.Scope),
                'Objective': request.data.get('Objective', original_policy.Objective),
                'Identifier': original_policy.Identifier,
                'PermanentTemporary': request.data.get('PermanentTemporary', original_policy.PermanentTemporary),
                'ActiveInactive': 'Inactive',
                'Reviewer': reviewer_name,  # Save UserName here
            }

            # Create new policy record
            new_policy = Policy.objects.create(**new_policy_data)

            # Find user ID for CreatedByName (for PolicyApproval)
            created_by_name = new_policy_data['CreatedByName']
            user_obj = Users.objects.filter(UserName=created_by_name).first()
            user_id = user_obj.UserId if user_obj else None

            # Prepare extracted data for PolicyApproval
            extracted_data = {
                'PolicyId': new_policy.PolicyId,
                'PolicyName': new_policy.PolicyName,
                'PolicyDescription': new_policy.PolicyDescription,
                'StartDate': new_policy.StartDate.isoformat() if isinstance(new_policy.StartDate, datetime.date) else new_policy.StartDate,
                'EndDate': new_policy.EndDate.isoformat() if isinstance(new_policy.EndDate, datetime.date) else new_policy.EndDate,
                'Department': new_policy.Department,
                'CreatedByName': new_policy.CreatedByName,
                'CreatedByDate': new_policy.CreatedByDate.isoformat() if isinstance(new_policy.CreatedByDate, datetime.date) else new_policy.CreatedByDate,
                'Applicability': new_policy.Applicability,
                'DocURL': new_policy.DocURL,
                'Scope': new_policy.Scope,
                'Objective': new_policy.Objective,
                'Identifier': new_policy.Identifier,
                'Status': new_policy.Status,
                'ActiveInactive': new_policy.ActiveInactive,
                'FrameworkId': new_policy.FrameworkId.FrameworkId,
                'policy_approval': {
                    'approved': None,
                    'remarks': ''
                },
                'subpolicies': []
            }

            # Create PolicyApproval record with ReviewerId as UserId
            PolicyApproval.objects.create(
                Identifier=new_policy.Identifier,
                ExtractedData=extracted_data,
                UserId=user_id,
                ReviewerId=reviewer_id,
                ApprovedNot=None,
                Version="u1"
            )

            # Get original PolicyVersion to link new version
            original_policy_version = PolicyVersion.objects.filter(
                PolicyId=original_policy,
                Version=str(original_policy.CurrentVersion)
            ).first()

            if not original_policy_version:
                return Response({
                    'error': f'No PolicyVersion found for PolicyId={original_policy.PolicyId} and Version={original_policy.CurrentVersion}. Data integrity issue.'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Create new PolicyVersion linked to previous
            policy_version = PolicyVersion(
                PolicyId=new_policy,
                Version=new_version,
                PolicyName=new_policy.PolicyName,
                CreatedBy=new_policy.CreatedByName,
                CreatedDate=new_policy.CreatedByDate,
                PreviousVersionId=original_policy_version.VersionId
            )
            policy_version.save()

            # Handle subpolicy customizations and new subpolicies (same as your original logic)
            subpolicy_customizations = {}
            subpolicies_to_exclude = []

            if 'subpolicies' in request.data:
                for sp_data in request.data.get('subpolicies', []):
                    if 'original_subpolicy_id' in sp_data:
                        sp_id = sp_data.get('original_subpolicy_id')
                        if sp_data.get('exclude', False):
                            subpolicies_to_exclude.append(sp_id)
                        else:
                            if 'Identifier' not in sp_data:
                                return Response({
                                    'error': 'Identifier is required for modified subpolicies',
                                    'subpolicy_id': sp_id
                                }, status=status.HTTP_400_BAD_REQUEST)
                            subpolicy_customizations[sp_id] = sp_data

            original_subpolicies = SubPolicy.objects.filter(PolicyId=original_policy)
            for original_subpolicy in original_subpolicies:
                if original_subpolicy.SubPolicyId in subpolicies_to_exclude:
                    continue
                custom_data = subpolicy_customizations.get(original_subpolicy.SubPolicyId, {})

                new_subpolicy_data = {
                    'PolicyId': new_policy,
                    'SubPolicyName': custom_data.get('SubPolicyName', original_subpolicy.SubPolicyName),
                    'CreatedByName': new_policy.CreatedByName,
                    'CreatedByDate': new_policy.CreatedByDate,
                    'Identifier': custom_data.get('Identifier', original_subpolicy.Identifier),
                    'Description': custom_data.get('Description', original_subpolicy.Description),
                    'Status': 'Under Review',
                    'PermanentTemporary': custom_data.get('PermanentTemporary', original_subpolicy.PermanentTemporary),
                    'Control': custom_data.get('Control', original_subpolicy.Control)
                }

                SubPolicy.objects.create(**new_subpolicy_data)

            # Add new subpolicies if any
            if 'new_subpolicies' in request.data:
                for new_subpolicy_data in request.data.get('new_subpolicies', []):
                    required_fields = ['SubPolicyName', 'Description', 'Identifier']
                    missing_fields = [field for field in required_fields if field not in new_subpolicy_data]
                    if missing_fields:
                        return Response({
                            'error': f'Missing required fields for new subpolicy: {", ".join(missing_fields)}'
                        }, status=status.HTTP_400_BAD_REQUEST)

                    subpolicy = new_subpolicy_data.copy()
                    subpolicy['PolicyId'] = new_policy
                    if 'CreatedByName' not in subpolicy:
                        subpolicy['CreatedByName'] = new_policy.CreatedByName
                    if 'CreatedByDate' not in subpolicy:
                        subpolicy['CreatedByDate'] = new_policy.CreatedByDate
                    if 'Status' not in subpolicy:
                        subpolicy['Status'] = 'Under Review'

                    SubPolicy.objects.create(**subpolicy)

            # Handle any new policies if specified
            created_policies = []
            if 'new_policies' in request.data:
                for new_policy_data in request.data.get('new_policies', []):
                    required_fields = ['PolicyName', 'PolicyDescription', 'Identifier']
                    missing_fields = [field for field in required_fields if field not in new_policy_data]
                    if missing_fields:
                        return Response({
                            'error': f'Missing required fields for new policy: {", ".join(missing_fields)}'
                        }, status=status.HTTP_400_BAD_REQUEST)

                    subpolicies_data = new_policy_data.pop('subpolicies', [])
                    policy_data = new_policy_data.copy()
                    policy_data['FrameworkId'] = original_policy.FrameworkId
                    policy_data['CurrentVersion'] = new_version
                    policy_data['Status'] = 'Under Review'
                    policy_data['ActiveInactive'] = 'Inactive'
                    if 'CreatedByName' not in policy_data:
                        policy_data['CreatedByName'] = original_policy.CreatedByName
                    if 'CreatedByDate' not in policy_data:
                        policy_data['CreatedByDate'] = datetime.date.today()

                    created_policy = Policy.objects.create(**policy_data)
                    created_policies.append(created_policy)

                    PolicyVersion.objects.create(
                        PolicyId=created_policy,
                        Version=new_version,
                        PolicyName=created_policy.PolicyName,
                        CreatedBy=created_policy.CreatedByName,
                        CreatedDate=created_policy.CreatedByDate,
                        PreviousVersionId=None
                    )

                    for subpolicy_data in subpolicies_data:
                        required_fields = ['SubPolicyName', 'Description', 'Identifier']
                        missing_fields = [field for field in required_fields if field not in subpolicy_data]
                        if missing_fields:
                            return Response({
                                'error': f'Missing required fields for subpolicy in new policy {created_policy.PolicyName}: {", ".join(missing_fields)}'
                            }, status=status.HTTP_400_BAD_REQUEST)

                        subpolicy = subpolicy_data.copy()
                        subpolicy['PolicyId'] = created_policy
                        if 'CreatedByName' not in subpolicy:
                            subpolicy['CreatedByName'] = created_policy.CreatedByName
                        if 'CreatedByDate' not in subpolicy:
                            subpolicy['CreatedByDate'] = created_policy.CreatedByDate
                        if 'Status' not in subpolicy:
                            subpolicy['Status'] = 'Under Review'

                        SubPolicy.objects.create(**subpolicy)

            # Prepare response
            response_data = {
                'message': 'New policy version created successfully',
                'PolicyId': new_policy.PolicyId,
                'PolicyName': new_policy.PolicyName,
                'PreviousVersion': current_version,
                'NewVersion': new_version,
                'FrameworkId': new_policy.FrameworkId.FrameworkId,
                'Identifier': new_policy.Identifier,
            }

            if created_policies:
                response_data['policies'] = [{
                    'PolicyId': p.PolicyId,
                    'PolicyName': p.PolicyName,
                    'Identifier': p.Identifier,
                    'Version': p.CurrentVersion
                } for p in created_policies]

            return Response(response_data, status=status.HTTP_201_CREATED)
    except Exception as e:
        error_info = {
            'error': str(e),
            'traceback': traceback.format_exc()
        }
        return Response({'error': 'Error creating new policy version', 'details': error_info}, status=status.HTTP_400_BAD_REQUEST)

"""
@api GET /api/frameworks/{framework_id}/export/
Exports all policies and their subpolicies for a specific framework to an Excel file in the following format:
Identifier, PolicyName (PolicyFamily), SubpolicyIdentifier, SubpolicyName, Control, Description

Example response:
Returns an Excel file as attachment
"""
@api_view(['GET'])
@permission_classes([AllowAny])
def export_policies_to_excel(request, framework_id):
    try:
        # Get the framework
        framework = get_object_or_404(Framework, FrameworkId=framework_id)
        
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from django.http import HttpResponse
        from datetime import datetime
        import re
        
        # Sanitize framework name for sheet title (remove invalid characters)
        def sanitize_sheet_name(name):
            # Remove or replace invalid characters for Excel sheet names
            # Excel sheet names cannot contain: \ / * ? : [ ]
            invalid_chars = r'[\\/*?:\[\]]'
            sanitized = re.sub(invalid_chars, '-', name)
            # Excel sheet names must be <= 31 characters
            return sanitized[:31].strip('-')  # Remove trailing dash if present
        
        def sanitize_filename(name):
            # Remove or replace invalid characters for file names
            invalid_chars = r'[<>:"/\\|?*]'
            return re.sub(invalid_chars, '-', name)
        
        # Create a new workbook and select the active sheet
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Sanitize and set the sheet title
        sheet_title = sanitize_sheet_name(f"{framework.FrameworkName} Policies")
        if not sheet_title:  # If all characters were invalid
            sheet_title = "Framework Policies"
        ws.title = sheet_title
        
        # Define headers
        headers = ['Identifier', 'PolicyFamily', 'SubpolicyIdentifier', 'SubpolicyName', 'Control', 'Description']
        
        # Style for headers
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        # Style for policy rows
        policy_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
        
        # Border styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Fetch all policies for this framework and their subpolicies
        policies = Policy.objects.filter(FrameworkId=framework).select_related('FrameworkId')
        
        row = 2  # Start from second row after headers
        for policy in policies:
            # Get subpolicies for this policy
            subpolicies = SubPolicy.objects.filter(PolicyId=policy)
            
            if not subpolicies:
                # If no subpolicies, write just the policy row
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border
                    if col <= 2:  # Only fill first two columns
                        cell.fill = policy_fill
                
                ws.cell(row=row, column=1, value=policy.Identifier)
                ws.cell(row=row, column=2, value=policy.PolicyName)
                row += 1
            else:
                # Store the starting row for this policy
                policy_start_row = row
                
                # Write policy with each subpolicy
                for subpolicy in subpolicies:
                    for col in range(1, 7):
                        cell = ws.cell(row=row, column=col)
                        cell.border = thin_border
                        if col <= 2:  # Only fill first two columns
                            cell.fill = policy_fill
                    
                    ws.cell(row=row, column=1, value=policy.Identifier)
                    ws.cell(row=row, column=2, value=policy.PolicyName)
                    ws.cell(row=row, column=3, value=subpolicy.Identifier)
                    ws.cell(row=row, column=4, value=subpolicy.SubPolicyName)
                    ws.cell(row=row, column=5, value=subpolicy.Control)
                    ws.cell(row=row, column=6, value=subpolicy.Description)
                    row += 1
                
                # Merge policy cells if there are multiple subpolicies
                if row - policy_start_row > 1:
                    ws.merge_cells(start_row=policy_start_row, start_column=1,
                                 end_row=row-1, end_column=1)
                    ws.merge_cells(start_row=policy_start_row, start_column=2,
                                 end_row=row-1, end_column=2)
                    
                    # Center the merged cells vertically
                    merged_cell1 = ws.cell(row=policy_start_row, column=1)
                    merged_cell2 = ws.cell(row=policy_start_row, column=2)
                    merged_cell1.alignment = Alignment(vertical='center')
                    merged_cell2.alignment = Alignment(vertical='center')
        
        # Add styling to the data cells and wrap text
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
            for cell in row:
                if not cell.alignment:  # Don't override merged cell alignment
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Adjust column widths
        column_widths = [20, 30, 20, 30, 40, 50]  # Preset widths for each column
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
        
        # Create response with Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Create a safe filename
        safe_framework_name = sanitize_filename(framework.FrameworkName)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f'{safe_framework_name}_policies_{timestamp}.xlsx'
        
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Save the workbook to the response
        wb.save(response)
        return response
        
    except Exception as e:
        return Response({
            'error': 'Error exporting policies to Excel',
            'details': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def policy_list(request):
    try:
        # Get query parameters
        framework_id = request.GET.get('framework_id')
        status = request.GET.get('status')
        active_inactive = request.GET.get('active_inactive')

        # Base queryset
        policies_query = Policy.objects.all()

        # Apply filters if provided
        if framework_id:
            policies_query = policies_query.filter(FrameworkId=framework_id)
        if status:
            policies_query = policies_query.filter(Status=status)
        if active_inactive:
            policies_query = policies_query.filter(ActiveInactive=active_inactive)

        # Get all policies
        policies = policies_query.select_related('FrameworkId')

        # Calculate summary counts
        summary_counts = {
            'active': Policy.objects.filter(ActiveInactive='Active').count(),
            'inactive': Policy.objects.filter(ActiveInactive='Inactive').count(),
            'approved': Policy.objects.filter(Status='Approved').count(),
            'rejected': Policy.objects.filter(Status='Rejected').count(),
            'under_review': Policy.objects.filter(Status='Under Review').count()
        }

        # Serialize policies
        serializer = PolicySerializer(policies, many=True)

        return Response({
            'policies': serializer.data,
            'summary_counts': summary_counts
        })

    except Exception as e:
        return Response({
            'error': 'Error fetching policies',
            'details': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def list_users(request):
    try:
        users = Users.objects.all()
        data = [{
            'UserId': user.UserId,
            'UserName': user.UserName
        } for user in users]
        return Response(data)
    except Exception as e:
        return Response({
            'error': 'Error fetching users',
            'details': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
@api_view(['GET'])
@permission_classes([AllowAny])
def get_framework_explorer_data(request):
    """
    API endpoint for the Framework Explorer page
    Returns frameworks with their status and counts of active/inactive policies
    """
    try:
        # Get all frameworks
        frameworks = Framework.objects.all()
       
        # Prepare response data with additional counts
        framework_data = []
        for fw in frameworks:
            # Count policies for this framework
            active_policies = Policy.objects.filter(
                FrameworkId=fw.FrameworkId,
                ActiveInactive='Active'
            ).count()
           
            inactive_policies = Policy.objects.filter(
                FrameworkId=fw.FrameworkId,
                ActiveInactive='Inactive'
            ).count()
           
            framework_data.append({
                'id': fw.FrameworkId,
                'name': fw.FrameworkName,
                'category': fw.Category or 'Uncategorized',
                'description': fw.FrameworkDescription,
                'status': fw.ActiveInactive,  # 'Active' or 'Inactive'
                'internalExternal': fw.InternalExternal or 'Internal',  # Add Internal/External field
                'active_policies_count': active_policies,
                'inactive_policies_count': inactive_policies
            })
       
        # Calculate summary counts
        active_frameworks = Framework.objects.filter(ActiveInactive='Active').count()
        inactive_frameworks = Framework.objects.filter(ActiveInactive='Inactive').count()
        active_policies = Policy.objects.filter(ActiveInactive='Active').count()
        inactive_policies = Policy.objects.filter(ActiveInactive='Inactive').count()
       
        return Response({
            'frameworks': framework_data,
            'summary': {
                'active_frameworks': active_frameworks,
                'inactive_frameworks': inactive_frameworks,
                'active_policies': active_policies,
                'inactive_policies': inactive_policies
            }
        })
       
    except Exception as e:
        return Response({
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
 
@api_view(['GET'])
@permission_classes([AllowAny])
def get_framework_policies(request, framework_id):
    """
    API endpoint for the Framework Policies page
    Returns policies for a specific framework
    """
    try:
        # Check if framework exists
        framework = get_object_or_404(Framework, FrameworkId=framework_id)
       
        # Get policies for this framework
        policies = Policy.objects.filter(FrameworkId=framework_id)
       
        # Prepare response data
        policy_data = []
        for policy in policies:
            policy_data.append({
                'id': policy.PolicyId,
                'name': policy.PolicyName,
                'category': policy.Department or 'General',
                'description': policy.PolicyDescription,
                'status': policy.ActiveInactive  # 'Active' or 'Inactive'
            })
       
        # Framework details
        framework_data = {
            'id': framework.FrameworkId,
            'name': framework.FrameworkName,
            'category': framework.Category,
            'description': framework.FrameworkDescription
        }
       
        return Response({
            'framework': framework_data,
            'policies': policy_data
        })
       
    except Exception as e:
        return Response({
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
 
@api_view(['POST'])
@permission_classes([AllowAny])
def toggle_framework_status(request, framework_id):
    """
    Toggle the ActiveInactive status of a framework
    When cascadeToApproved=True, also update the status of all approved policies
    but leave their subpolicies status unchanged
    When activating a framework, all other related versions (connected through PreviousVersionId chain)
    are set to inactive
    """
    try:
        framework = get_object_or_404(Framework, FrameworkId=framework_id)
       
        # Toggle status
        new_status = 'Inactive' if framework.ActiveInactive == 'Active' else 'Active'
        framework.ActiveInactive = new_status
        framework.save()
        
        # When activating a framework, set all other related versions to inactive
        other_versions_deactivated = 0
        if new_status == 'Active':
            # Find all versions connected to this framework through the PreviousVersionId chain
            related_version_ids = set()
            
            # First, get all framework versions associated with this framework
            try:
                framework_versions = FrameworkVersion.objects.filter(FrameworkId=framework)
                if framework_versions.exists():
                    # For each version, find all related versions through PreviousVersionId chain
                    for version in framework_versions:
                        # Start with this version's ID
                        to_process = [version.VersionId]
                        processed = set()
                        
                        # Follow chain forward and backward
                        while to_process:
                            current_id = to_process.pop()
                            if current_id in processed:
                                continue
                            
                            processed.add(current_id)
                            
                            # Find versions where this is PreviousVersionId (forward)
                            next_versions = FrameworkVersion.objects.filter(PreviousVersionId=current_id)
                            for next_ver in next_versions:
                                related_version_ids.add(next_ver.FrameworkId.FrameworkId)
                                if next_ver.VersionId not in processed:
                                    to_process.append(next_ver.VersionId)
                            
                            # Find version that this version points to (backward)
                            try:
                                current_version = FrameworkVersion.objects.get(VersionId=current_id)
                                if current_version.PreviousVersionId and current_version.PreviousVersionId not in processed:
                                    prev_version = FrameworkVersion.objects.get(VersionId=current_version.PreviousVersionId)
                                    related_version_ids.add(prev_version.FrameworkId.FrameworkId)
                                    to_process.append(current_version.PreviousVersionId)
                            except FrameworkVersion.DoesNotExist:
                                pass
                    
                    # Remove the current framework from the related list
                    if framework.FrameworkId in related_version_ids:
                        related_version_ids.remove(framework.FrameworkId)
                    
                    # Set all related frameworks to inactive
                    if related_version_ids:
                        related_frameworks = Framework.objects.filter(FrameworkId__in=related_version_ids)
                        for related_framework in related_frameworks:
                            if related_framework.ActiveInactive == 'Active':
                                related_framework.ActiveInactive = 'Inactive'
                                related_framework.save()
                                other_versions_deactivated += 1
            except Exception as e:
                print(f"Error finding related framework versions: {str(e)}")
        
        policies_affected = 0
        subpolicies_affected = 0
        
        # Check if we should cascade to approved policies
        cascade_to_approved = request.data.get('cascadeToApproved', False)
        if cascade_to_approved:
            # Get all approved policies for this framework
            approved_policies = Policy.objects.filter(
                FrameworkId=framework,
                Status='Approved'
            )
            
            # Update their status to match the framework
            for policy in approved_policies:
                policy.ActiveInactive = new_status
                policy.save()
                policies_affected += 1
                
                # Count subpolicies but don't change their status
                subpolicies_count = SubPolicy.objects.filter(PolicyId=policy).count()
                subpolicies_affected += subpolicies_count
       
        return Response({
            'id': framework.FrameworkId,
            'status': new_status,
            'message': f'Framework status updated to {new_status}',
            'policies_affected': policies_affected,
            'subpolicies_affected': subpolicies_affected,
            'other_versions_deactivated': other_versions_deactivated
        })
       
    except Exception as e:
        return Response({
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([AllowAny])
def toggle_policy_status(request, policy_id):
    """
    Toggle the ActiveInactive status of a policy
    When cascadeSubpolicies=True, we don't change subpolicy Status but still count them
    When activating a policy, all other related versions (connected through PreviousVersionId chain)
    are set to inactive
    """
    try:
        policy = get_object_or_404(Policy, PolicyId=policy_id)
       
        # Toggle status
        new_status = 'Inactive' if policy.ActiveInactive == 'Active' else 'Active'
        policy.ActiveInactive = new_status
        policy.save()
        
        # When activating a policy, set all other related versions to inactive
        other_versions_deactivated = 0
        if new_status == 'Active':
            # Find all versions connected to this policy through the PreviousVersionId chain
            related_policy_ids = set()
            
            # First, get all policy versions associated with this policy
            try:
                policy_versions = PolicyVersion.objects.filter(PolicyId=policy)
                if policy_versions.exists():
                    # For each version, find all related versions through PreviousVersionId chain
                    for version in policy_versions:
                        # Start with this version's ID
                        to_process = [version.VersionId]
                        processed = set()
                        
                        # Follow chain forward and backward
                        while to_process:
                            current_id = to_process.pop()
                            if current_id in processed:
                                continue
                            
                            processed.add(current_id)
                            
                            # Find versions where this is PreviousVersionId (forward)
                            next_versions = PolicyVersion.objects.filter(PreviousVersionId=current_id)
                            for next_ver in next_versions:
                                related_policy_ids.add(next_ver.PolicyId_id)
                                if next_ver.VersionId not in processed:
                                    to_process.append(next_ver.VersionId)
                            
                            # Find version that this version points to (backward)
                            try:
                                current_version = PolicyVersion.objects.get(VersionId=current_id)
                                if current_version.PreviousVersionId and current_version.PreviousVersionId not in processed:
                                    prev_version = PolicyVersion.objects.get(VersionId=current_version.PreviousVersionId)
                                    related_policy_ids.add(prev_version.PolicyId_id)
                                    to_process.append(current_version.PreviousVersionId)
                            except PolicyVersion.DoesNotExist:
                                pass
                    
                    # Remove the current policy from the related list
                    if policy.PolicyId in related_policy_ids:
                        related_policy_ids.remove(policy.PolicyId)
                    
                    # Set all related policies to inactive
                    if related_policy_ids:
                        related_policies = Policy.objects.filter(PolicyId__in=related_policy_ids)
                        for related_policy in related_policies:
                            if related_policy.ActiveInactive == 'Active':
                                related_policy.ActiveInactive = 'Inactive'
                                related_policy.save()
                                other_versions_deactivated += 1
            except Exception as e:
                print(f"Error finding related policy versions: {str(e)}")
        
        subpolicies_affected = 0
        
        # Count subpolicies but don't change their status
        cascade_to_subpolicies = request.data.get('cascadeSubpolicies', False)
        if cascade_to_subpolicies:
            # Get count of all subpolicies for this policy
            subpolicies_affected = SubPolicy.objects.filter(PolicyId=policy).count()
       
        return Response({
            'id': policy.PolicyId,
            'status': new_status,
            'message': f'Policy status updated to {new_status}',
            'subpolicies_affected': subpolicies_affected,
            'other_versions_deactivated': other_versions_deactivated
        })
       
    except Exception as e:
        return Response({
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
 
@api_view(['GET'])
@permission_classes([AllowAny])
def get_framework_details(request, framework_id):
    """
    API endpoint for detailed framework information
    Returns all details of a framework regardless of status
    """
    try:
        # Get framework by ID
        framework = get_object_or_404(Framework, FrameworkId=framework_id)
       
        # Create response data
        response_data = {
            'FrameworkId': framework.FrameworkId,
            'FrameworkName': framework.FrameworkName,
            'CurrentVersion': framework.CurrentVersion,
            'FrameworkDescription': framework.FrameworkDescription,
            'EffectiveDate': framework.EffectiveDate,
            'CreatedByName': framework.CreatedByName,
            'CreatedByDate': framework.CreatedByDate,
            'Category': framework.Category,
            'DocURL': framework.DocURL,
            'Identifier': framework.Identifier,
            'StartDate': framework.StartDate,
            'EndDate': framework.EndDate,
            'Status': framework.Status,
            'ActiveInactive': framework.ActiveInactive
        }
       
        return Response(response_data)
       
    except Exception as e:
        return Response({
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
 
@api_view(['GET'])
@permission_classes([AllowAny])
def get_policy_details(request, policy_id):
    """
    API endpoint for detailed policy information
    Returns all details of a policy regardless of status
    """
    try:
        # Get policy by ID
        policy = get_object_or_404(Policy, PolicyId=policy_id)
       
        # Get all subpolicies for this policy
        subpolicies = SubPolicy.objects.filter(PolicyId=policy)
        subpolicy_data = []
       
        for subpolicy in subpolicies:
            subpolicy_data.append({
                'SubPolicyId': subpolicy.SubPolicyId,
                'SubPolicyName': subpolicy.SubPolicyName,
                'CreatedByName': subpolicy.CreatedByName,
                'CreatedByDate': subpolicy.CreatedByDate,
                'Identifier': subpolicy.Identifier,
                'Description': subpolicy.Description,
                'Status': subpolicy.Status,
                'PermanentTemporary': subpolicy.PermanentTemporary,
                'Control': subpolicy.Control
            })
       
        # Create response data
        response_data = {
            'PolicyId': policy.PolicyId,
            'PolicyName': policy.PolicyName,
            'PolicyDescription': policy.PolicyDescription,
            'CurrentVersion': policy.CurrentVersion,
            'StartDate': policy.StartDate,
            'EndDate': policy.EndDate,
            'Department': policy.Department,
            'CreatedByName': policy.CreatedByName,
            'CreatedByDate': policy.CreatedByDate,
            'Applicability': policy.Applicability,
            'DocURL': policy.DocURL,
            'Scope': policy.Scope,
            'Objective': policy.Objective,
            'Identifier': policy.Identifier,
            'PermanentTemporary': policy.PermanentTemporary,
            'Status': policy.Status,
            'ActiveInactive': policy.ActiveInactive,
            'subpolicies': subpolicy_data
        }
       
        return Response(response_data)
       
    except Exception as e:
        return Response({
            'error': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

#all policies code
@api_view(['GET'])
@permission_classes([AllowAny])
def all_policies_get_frameworks(request):
    """
    API endpoint to get all frameworks for AllPolicies.vue component.
    """
    try:
        frameworks = Framework.objects.all()
        
        frameworks_data = []
        for framework in frameworks:
            framework_data = {
                'id': framework.FrameworkId,
                'name': framework.FrameworkName,
                'category': framework.Category,
                'status': framework.ActiveInactive,
                'description': framework.FrameworkDescription,
                'versions': []
            }
            
            # Get versions for this framework
            versions = FrameworkVersion.objects.filter(FrameworkId=framework)
            version_data = []
            for version in versions:
                version_data.append({
                    'id': version.VersionId,
                    'name': f"v{version.Version}",
                    'version': version.Version
                })
            
            framework_data['versions'] = version_data
            frameworks_data.append(framework_data)
            
        return Response(frameworks_data)
    
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def all_policies_get_framework_version_policies(request, version_id):
    """
    API endpoint to get all policies for a specific framework version for AllPolicies.vue component.
    """
    try:
        # Get the framework version
        framework_version = get_object_or_404(FrameworkVersion, VersionId=version_id)
        framework = framework_version.FrameworkId
        
        # Get policies for this framework
        policies = Policy.objects.filter(
            FrameworkId=framework, 
            CurrentVersion=framework_version.Version
        )
        
        policies_data = []
        for policy in policies:
            policy_data = {
                'id': policy.PolicyId,
                'name': policy.PolicyName,
                'category': policy.Department,
                'status': policy.Status,
                'description': policy.PolicyDescription,
                'versions': []
            }
            
            # Get versions for this policy
            policy_versions = PolicyVersion.objects.filter(PolicyId=policy)
            versions_data = []
            for version in policy_versions:
                versions_data.append({
                    'id': version.VersionId,
                    'name': f"v{version.Version}",
                    'version': version.Version
                })
            
            policy_data['versions'] = versions_data
            policies_data.append(policy_data)
            
        return Response(policies_data)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def all_policies_get_policies(request):
    """
    API endpoint to get all policies for AllPolicies.vue component.
    """
    try:
        # Optional framework filter
        framework_id = request.GET.get('framework_id')
        
        # Start with all policies
        policies_query = Policy.objects.all()
        
        # Apply framework filter if provided
        if framework_id:
            policies_query = policies_query.filter(FrameworkId_id=framework_id)
        
        policies_data = []
        for policy in policies_query:
            policy_data = {
                'id': policy.PolicyId,
                'name': policy.PolicyName,
                'category': policy.Department,
                'status': policy.Status,
                'description': policy.PolicyDescription,
                'versions': []
            }
            
            # Get versions for this policy
            policy_versions = PolicyVersion.objects.filter(PolicyId=policy)
            versions_data = []
            for version in policy_versions:
                versions_data.append({
                    'id': version.VersionId,
                    'name': f"v{version.Version}",
                    'version': version.Version
                })
            
            policy_data['versions'] = versions_data
            policies_data.append(policy_data)
            
        return Response(policies_data)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def all_policies_get_policy_versions(request, policy_id):
    """
    API endpoint to get all versions of a specific policy for AllPolicies.vue component.
    Implements a dedicated version that handles version chains through PreviousVersionId.
    """
    try:
        print(f"Request received for policy versions, policy_id: {policy_id}, type: {type(policy_id)}")
        
        # Ensure we have a valid integer ID
        try:
            policy_id = int(policy_id)
        except (ValueError, TypeError):
            return Response({'error': f'Invalid policy ID format: {policy_id}'}, 
                           status=status.HTTP_400_BAD_REQUEST)
        
        # Get the base policy
        try:
            policy = Policy.objects.get(PolicyId=policy_id)
            print(f"Found policy: {policy.PolicyName} (ID: {policy.PolicyId})")
        except Policy.DoesNotExist:
            print(f"Policy with ID {policy_id} not found")
            return Response({'error': f'Policy with ID {policy_id} not found'}, 
                           status=status.HTTP_404_NOT_FOUND)
        
        # Get the direct policy version
        try:
            direct_version = PolicyVersion.objects.get(PolicyId=policy)
            print(f"Found direct policy version: {direct_version.VersionId}")
        except PolicyVersion.DoesNotExist:
            print(f"No policy version found for policy ID {policy_id}")
            return Response({'error': f'No version found for policy with ID {policy_id}'}, 
                           status=status.HTTP_404_NOT_FOUND)
        except PolicyVersion.MultipleObjectsReturned:
            # If there are multiple versions, get all of them
            direct_versions = list(PolicyVersion.objects.filter(PolicyId=policy))
            print(f"Found {len(direct_versions)} direct versions for policy {policy_id}")
            direct_version = direct_versions[0]  # Just use the first one for starting the chain
        
        # Start building version chain
        all_versions = {}
        visited = set()
        to_process = [direct_version.VersionId]
        
        # Find all versions in the chain
        while to_process:
            current_id = to_process.pop(0)
            
            if current_id in visited:
                continue
                
            visited.add(current_id)
            
            try:
                current_version = PolicyVersion.objects.get(VersionId=current_id)
                all_versions[current_id] = current_version
                
                # Follow PreviousVersionId chain backward
                if current_version.PreviousVersionId and current_version.PreviousVersionId not in visited:
                    to_process.append(current_version.PreviousVersionId)
                    
                # Find versions that reference this one as their previous version
                next_versions = PolicyVersion.objects.filter(PreviousVersionId=current_id)
                for next_ver in next_versions:
                    if next_ver.VersionId not in visited:
                        to_process.append(next_ver.VersionId)
            except PolicyVersion.DoesNotExist:
                print(f"Version with ID {current_id} not found")
                continue
        
        versions_data = []
        for version_id, version in all_versions.items():
            try:
                # Get the policy this version belongs to
                version_policy = version.PolicyId
                
                # Count subpolicies for this policy
                subpolicy_count = SubPolicy.objects.filter(PolicyId=version_policy).count()
                
                # Get previous version details if available
                previous_version = None
                if version.PreviousVersionId:
                    try:
                        previous_version = PolicyVersion.objects.get(VersionId=version.PreviousVersionId)
                    except PolicyVersion.DoesNotExist:
                        pass
                
                # Create a descriptive name
                formatted_name = f"{version.PolicyName} v{version.Version}" if version.PolicyName else f"{version_policy.PolicyName} v{version.Version}"
                
                version_data = {
                    'id': version.VersionId,
                    'policy_id': version_policy.PolicyId,
                    'name': formatted_name,
                    'version': version.Version,
                    'category': version_policy.Department or 'General',
                    'status': version_policy.Status or 'Unknown',
                    'description': version_policy.PolicyDescription or '',
                    'created_date': version.CreatedDate,
                    'created_by': version.CreatedBy,
                    'subpolicy_count': subpolicy_count,
                    'previous_version_id': version.PreviousVersionId,
                    'previous_version_name': previous_version.PolicyName + f" v{previous_version.Version}" if previous_version else None
                }
                versions_data.append(version_data)
                print(f"Added version: {version.VersionId} - {formatted_name}, Previous: {version.PreviousVersionId}")
            except Exception as e:
                print(f"Error processing version {version_id}: {str(e)}")
                # Continue to next version
        
        # Sort versions by version number (descending)
        versions_data.sort(key=lambda x: float(x['version']), reverse=True)

        
        
        print(f"Returning {len(versions_data)} policy versions")
        return Response(versions_data)
        
    except Exception as e:
        import traceback
        print(f"Error in all_policies_get_policy_versions: {str(e)}")
        traceback.print_exc()
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def all_policies_get_subpolicies(request):
    """
    API endpoint to get all subpolicies for AllPolicies.vue component.
    """
    try:
        print("Request received for all subpolicies")
        
        # Optional framework filter
        framework_id = request.GET.get('framework_id')
        print(f"Framework filter: {framework_id}")
        
        # Start with all subpolicies
        subpolicies_query = SubPolicy.objects.all()
        
        # If framework filter is provided, filter through policies
        if framework_id:
            try:
                policy_ids = Policy.objects.filter(FrameworkId_id=framework_id).values_list('PolicyId', flat=True)
                print(f"Found {len(policy_ids)} policies for framework {framework_id}")
                subpolicies_query = subpolicies_query.filter(PolicyId_id__in=policy_ids)
            except Exception as e:
                print(f"Error filtering by framework: {str(e)}")
                # Continue with all subpolicies if framework filtering fails
        
        print(f"Found {subpolicies_query.count()} subpolicies")
        
        subpolicies_data = []
        for subpolicy in subpolicies_query:
            try:
                # Get the policy this subpolicy belongs to
                try:
                    policy = Policy.objects.get(PolicyId=subpolicy.PolicyId_id)
                    policy_name = policy.PolicyName
                    department = policy.Department
                except Policy.DoesNotExist:
                    print(f"Policy with ID {subpolicy.PolicyId_id} not found for subpolicy {subpolicy.SubPolicyId}")
                    policy_name = "Unknown Policy"
                    department = "Unknown"
                
                subpolicy_data = {
                    'id': subpolicy.SubPolicyId,
                    'name': subpolicy.SubPolicyName,
                    'category': department or 'General',
                    'status': subpolicy.Status or 'Unknown',
                    'description': subpolicy.Description or '',
                    'control': subpolicy.Control or '',
                    'identifier': subpolicy.Identifier,
                    'permanent_temporary': subpolicy.PermanentTemporary,
                    'policy_id': subpolicy.PolicyId_id,
                    'policy_name': policy_name,
                    'created_by': subpolicy.CreatedByName,
                    'created_date': subpolicy.CreatedByDate
                }
                subpolicies_data.append(subpolicy_data)
                print(f"Added subpolicy: {subpolicy.SubPolicyId} - {subpolicy.SubPolicyName}")
            except Exception as e:
                print(f"Error processing subpolicy {subpolicy.SubPolicyId}: {str(e)}")
                # Continue to next subpolicy
        
        print(f"Returning {len(subpolicies_data)} subpolicies")
        return Response(subpolicies_data)
        
    except Exception as e:
        import traceback
        print(f"Error in all_policies_get_subpolicies: {str(e)}")
        traceback.print_exc()
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def all_policies_get_subpolicy_details(request, subpolicy_id):
    """
    API endpoint to get details of a specific subpolicy for AllPolicies.vue component.
    """
    try:
        subpolicy = get_object_or_404(SubPolicy, SubPolicyId=subpolicy_id)
        policy = subpolicy.PolicyId
        
        subpolicy_data = {
            'id': subpolicy.SubPolicyId,
            'name': subpolicy.SubPolicyName,
            'category': policy.Department,
            'status': subpolicy.Status,
            'description': subpolicy.Description,
            'control': subpolicy.Control,
            'identifier': subpolicy.Identifier,
            'permanent_temporary': subpolicy.PermanentTemporary,
            'policy_id': policy.PolicyId,
            'policy_name': policy.PolicyName
        }
        
        return Response(subpolicy_data)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def all_policies_get_framework_versions(request, framework_id):
    """
    API endpoint to get all versions of a specific framework for AllPolicies.vue component.
    Implements a dedicated version that handles version chains through PreviousVersionId.
    """
    try:
        print(f"Request received for framework versions, framework_id: {framework_id}, type: {type(framework_id)}")
        
        # Ensure we have a valid integer ID
        try:
            framework_id = int(framework_id)
        except (ValueError, TypeError):
            return Response({'error': f'Invalid framework ID format: {framework_id}'}, 
                           status=status.HTTP_400_BAD_REQUEST)
        
        # Get the base framework
        try:
            framework = Framework.objects.get(FrameworkId=framework_id)
            print(f"Found framework: {framework.FrameworkName} (ID: {framework.FrameworkId})")
        except Framework.DoesNotExist:
            print(f"Framework with ID {framework_id} not found")
            return Response({'error': f'Framework with ID {framework_id} not found'}, 
                           status=status.HTTP_404_NOT_FOUND)
        
        # Get direct versions that belong to this framework
        direct_versions = list(FrameworkVersion.objects.filter(FrameworkId=framework).order_by('-Version'))
        print(f"Found {len(direct_versions)} direct versions for framework {framework_id}")
        
        # Create a dictionary to track all versions by VersionId
        all_versions = {v.VersionId: v for v in direct_versions}
        
        # Create a queue to process versions and follow PreviousVersionId links
        to_process = [v.VersionId for v in direct_versions]
        
        # Process the version chain by following PreviousVersionId links
        while to_process:
            current_id = to_process.pop(0)
            
            # Find versions that reference this one as their previous version
            linked_versions = FrameworkVersion.objects.filter(PreviousVersionId=current_id)
            print(f"Found {len(linked_versions)} linked versions for version ID {current_id}")
            
            for linked in linked_versions:
                if linked.VersionId not in all_versions:
                    # Add newly discovered version to our collection
                    all_versions[linked.VersionId] = linked
                    to_process.append(linked.VersionId)
        
        versions_data = []
        for version_id, version in all_versions.items():
            try:
                # Get the framework this version belongs to
                version_framework = version.FrameworkId
                
                # Count policies for this framework (without filtering by version)
                # This gets all policies associated with this framework regardless of version
                policy_count = Policy.objects.filter(
                    FrameworkId=version_framework
                ).count()
                
                print(f"Found {policy_count} policies for framework {version_framework.FrameworkId}")
                
                # Get previous version details if available
                previous_version = None
                if version.PreviousVersionId:
                    try:
                        previous_version = FrameworkVersion.objects.get(VersionId=version.PreviousVersionId)
                    except FrameworkVersion.DoesNotExist:
                        pass
                
                # Create a more descriptive name using the FrameworkName from the database
                # and appending the version number like v1.0, v2.0, etc.
                formatted_name = f"{version.FrameworkName} v{version.Version}"
                
                version_data = {
                    'id': version.VersionId,
                    'name': formatted_name,
                    'version': version.Version,
                    'category': version_framework.Category or 'General',
                    'status': version_framework.ActiveInactive or 'Unknown',
                    'description': version_framework.FrameworkDescription or '',
                    'created_date': version.CreatedDate,
                    'created_by': version.CreatedBy,
                    'policy_count': policy_count,
                    'previous_version_id': version.PreviousVersionId,
                    'previous_version_name': previous_version.FrameworkName + f" v{previous_version.Version}" if previous_version else None,
                    'framework_id': version_framework.FrameworkId
                }
                versions_data.append(version_data)
                print(f"Added version: {version.VersionId} - {formatted_name}, Previous: {version.PreviousVersionId}")
            except Exception as e:
                print(f"Error processing version {version_id}: {str(e)}")
                # Continue to next version
        
        # Sort versions by version number (descending)
        versions_data.sort(key=lambda x: float(x['version']), reverse=True)
        
        print(f"Returning {len(versions_data)} versions")
        return Response(versions_data)
        
    except Exception as e:
        import traceback
        print(f"Error in all_policies_get_framework_versions: {str(e)}")
        traceback.print_exc()
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def all_policies_get_policy_version_subpolicies(request, version_id):
    """
    API endpoint to get all subpolicies for a specific policy version for AllPolicies.vue component.
    Implements a dedicated version instead of using the existing get_policy_version_subpolicies function.
    """
    try:
        print(f"Request received for policy version subpolicies, version_id: {version_id}, type: {type(version_id)}")
        
        # Ensure we have a valid integer ID
        try:
            version_id = int(version_id)
        except (ValueError, TypeError):
            print(f"Invalid version ID format: {version_id}")
            return Response({'error': f'Invalid version ID format: {version_id}'}, 
                           status=status.HTTP_400_BAD_REQUEST)
        
        # Get the policy version
        try:
            policy_version = PolicyVersion.objects.get(VersionId=version_id)
            print(f"Found policy version: {policy_version.VersionId} for policy {policy_version.PolicyId_id}")
        except PolicyVersion.DoesNotExist:
            print(f"Policy version with ID {version_id} not found")
            return Response({'error': f'Policy version with ID {version_id} not found'}, 
                           status=status.HTTP_404_NOT_FOUND)
        
        # Get the policy this version belongs to
        try:
            policy = Policy.objects.get(PolicyId=policy_version.PolicyId_id)
            print(f"Found policy: {policy.PolicyName} (ID: {policy.PolicyId})")
        except Policy.DoesNotExist:
            print(f"Policy with ID {policy_version.PolicyId_id} not found")
            return Response({'error': f'Policy with ID {policy_version.PolicyId_id} not found'}, 
                           status=status.HTTP_404_NOT_FOUND)
        
        # Get subpolicies for this policy
        subpolicies = SubPolicy.objects.filter(PolicyId=policy)
        print(f"Found {len(subpolicies)} subpolicies for policy {policy.PolicyId}")
        
        subpolicies_data = []
        for subpolicy in subpolicies:
            try:
                subpolicy_data = {
                    'id': subpolicy.SubPolicyId,
                    'name': subpolicy.SubPolicyName,
                    'category': policy.Department or 'General',
                    'status': subpolicy.Status or 'Unknown',
                    'description': subpolicy.Description or '',
                    'control': subpolicy.Control or '',
                    'identifier': subpolicy.Identifier,
                    'permanent_temporary': subpolicy.PermanentTemporary,
                    'policy_id': policy.PolicyId,
                    'policy_name': policy.PolicyName,
                    'created_by': subpolicy.CreatedByName,
                    'created_date': subpolicy.CreatedByDate
                }
                subpolicies_data.append(subpolicy_data)
                print(f"Added subpolicy: {subpolicy.SubPolicyId} - {subpolicy.SubPolicyName}")
            except Exception as e:
                print(f"Error processing subpolicy {subpolicy.SubPolicyId}: {str(e)}")
                # Continue to next subpolicy
        
        print(f"Returning {len(subpolicies_data)} subpolicies")
        return Response(subpolicies_data)
        
    except Exception as e:
        import traceback
        print(f"Error in all_policies_get_policy_version_subpolicies: {str(e)}")
        traceback.print_exc()
        return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_policy_dashboard_summary(request):
    total_policies = Policy.objects.count()
    total_subpolicies = SubPolicy.objects.count()
    active_policies = Policy.objects.filter(ActiveInactive='Active').count()
    inactive_policies = Policy.objects.filter(ActiveInactive='Inactive').count()
    approved_policies = PolicyApproval.objects.filter(ApprovedNot=1).count()
    total_approvals = PolicyApproval.objects.count()
    approval_rate = (approved_policies / total_approvals) * 100 if total_approvals else 0

    # Get all policies with their details
    policies = Policy.objects.all().values(
        'PolicyId', 'PolicyName', 'Department', 'Status', 
        'Applicability', 'CurrentVersion', 'ActiveInactive',
        'PermanentTemporary', 'CreatedByDate'
    )

    return Response({
        'total_policies': total_policies,
        'total_subpolicies': total_subpolicies,
        'active_policies': active_policies,
        'inactive_policies': inactive_policies,
        'approval_rate': round(approval_rate, 2),
        'policies': list(policies)
    })

@api_view(['GET'])
@permission_classes([AllowAny])
def get_policy_status_distribution(request):
    status_counts = Policy.objects.values('Status').annotate(count=Count('Status'))
    return Response(status_counts)

@api_view(['GET'])
@permission_classes([AllowAny])
def get_reviewer_workload(request):
    reviewer_counts = Policy.objects.values('Reviewer').annotate(count=Count('Reviewer')).order_by('-count')
    return Response(reviewer_counts)

from django.utils import timezone
from datetime import timedelta

@api_view(['GET'])
@permission_classes([AllowAny])
def get_recent_policy_activity(request):
    one_week_ago = timezone.now().date() - timedelta(days=7)
    recent_policies = Policy.objects.filter(CreatedByDate__gte=one_week_ago).order_by('-CreatedByDate')[:10]
    return Response([
        {
            'PolicyName': p.PolicyName,
            'CreatedBy': p.CreatedByName,
            'CreatedDate': p.CreatedByDate
        } for p in recent_policies
    ])

from django.db.models import F, ExpressionWrapper, DurationField

@api_view(['GET'])
@permission_classes([AllowAny])
def get_avg_policy_approval_time(request):
    # Get all approved policies with approval dates
    approved_policies = PolicyApproval.objects.filter(
        ApprovedNot=1,
        ApprovedDate__isnull=False
    )
    
    if not approved_policies:
        return Response({'average_days': 0})

    # Get the first and last approval for each policy
    policy_approvals = {}
    for approval in approved_policies:
        if approval.Identifier not in policy_approvals:
            policy_approvals[approval.Identifier] = {
                'first': approval,
                'last': approval
            }
        else:
            if approval.ApprovalId < policy_approvals[approval.Identifier]['first'].ApprovalId:
                policy_approvals[approval.Identifier]['first'] = approval
            if approval.ApprovalId > policy_approvals[approval.Identifier]['last'].ApprovalId:
                policy_approvals[approval.Identifier]['last'] = approval

    # Calculate average days between first submission and approval
    total_days = 0
    count = 0
    for approvals in policy_approvals.values():
        if approvals['first'].ApprovedDate and approvals['last'].ApprovedDate:
            days = (approvals['last'].ApprovedDate - approvals['first'].ApprovedDate).days
            if days >= 0:  # Only count positive durations
                total_days += days
                count += 1

    avg_days = total_days / count if count > 0 else 0
    return Response({'average_days': round(avg_days, 2)})

@api_view(['GET'])
@permission_classes([AllowAny])
def get_policy_analytics(request):
    try:
        x_axis = request.GET.get('x_axis', 'time')
        y_axis = request.GET.get('y_axis', 'count')
        
        # Initialize base queryset
        if x_axis == 'subpolicy':
            queryset = SubPolicy.objects.all()
            base_model = 'subpolicy'
        elif x_axis == 'framework':
            queryset = Framework.objects.all()
            base_model = 'framework'
        else:
            queryset = Policy.objects.all()
            base_model = 'policy'

        # Select base queryset based on x-axis and y-axis combination
        if y_axis == 'framework_policies' and x_axis == 'time':
            # Count policies created on each date
            queryset = Policy.objects.values(
                'CreatedByDate'
            ).annotate(
                label=F('CreatedByDate'),
                value=Count('PolicyId')
            ).order_by('CreatedByDate')
            
            # Format the dates
            data = list(queryset)
            for item in data:
                item['label'] = item['label'].strftime('%Y-%m-%d') if item['label'] else None
            
            return Response(data)
        
        # Group by X-axis
        if x_axis == 'status':
            queryset = queryset.values(
                'Status'
            ).annotate(
                label=F('Status'),
            )
        elif x_axis == 'policy':
            queryset = queryset.values(
                'PolicyId', 'PolicyName'
            ).annotate(
                label=F('PolicyName'),
            )
        elif x_axis == 'subpolicy':
            queryset = queryset.values(
                'SubPolicyId', 'SubPolicyName'
            ).annotate(
                label=F('SubPolicyName'),
            )
        elif x_axis == 'time':
            date_field = {
                'framework': 'CreatedByDate',
                'policy': 'CreatedByDate',
                'subpolicy': 'CreatedByDate'
            }[base_model]
            queryset = queryset.values(
                date_field
            ).annotate(
                label=F(date_field),
            ).order_by(date_field)
        elif x_axis == 'framework':
            queryset = queryset.values(
                'FrameworkName'
            ).annotate(
                label=F('FrameworkName'),
            )
        
        # Apply Y-axis aggregation
        if y_axis == 'version':
            if base_model == 'framework':
                # Get all framework versions
                framework_versions = FrameworkVersion.objects.values(
                    'FrameworkId__Identifier'  # Group by framework identifier
                ).annotate(
                    version_count=Count('VersionId', distinct=True)  # Count distinct versions
                )
                
                # Create a mapping of framework identifier to version count
                version_counts = {fv['FrameworkId__Identifier']: fv['version_count'] for fv in framework_versions}
                
                # Get all frameworks first
                frameworks = Framework.objects.all()
                framework_id_to_identifier = {f.FrameworkId: f.Identifier for f in frameworks}
                
                # Add version count to each item in queryset
                data = list(queryset)
                for item in data:
                    if x_axis == 'framework':
                        framework_id = item['FrameworkId']
                    else:
                        # For other x_axis types (like department), we need to aggregate versions
                        # Get all frameworks matching the current group
                        if x_axis == 'department':
                            group_frameworks = frameworks.filter(Department=item['Department'])
                        elif x_axis == 'status':
                            group_frameworks = frameworks.filter(Status=item['Status'])
                        elif x_axis == 'applicability':
                            group_frameworks = frameworks.filter(Applicability=item['Applicability'])
                        else:
                            group_frameworks = frameworks
                            
                        # Sum up versions for all frameworks in this group
                        total_versions = 0
                        for framework in group_frameworks:
                            total_versions += version_counts.get(framework.Identifier, 1)
                        item['value'] = total_versions
                        continue
                    
                    # For framework x_axis, use the direct mapping
                    identifier = framework_id_to_identifier.get(framework_id)
                    item['value'] = version_counts.get(identifier, 1) if identifier else 1
                
                return Response(data)
                
            elif base_model == 'policy':
                # Count versions by following PreviousVersionId chain
                policy_versions = PolicyVersion.objects.values('PolicyId').annotate(
                    version_count=Count('VersionId', distinct=True)
                )
                version_counts = {pv['PolicyId']: pv['version_count'] for pv in policy_versions}
                
                # Add version count to each policy
                data = list(queryset)
                for item in data:
                    item['value'] = version_counts.get(item.get('PolicyId'), 0)
                return Response(data)
            else:
                # SubPolicies don't have versions
                queryset = queryset.annotate(value=Value(0))
        elif y_axis == 'activeInactive':
            if base_model == 'framework':
                # For frameworks, use ActiveInactive field
                queryset = queryset.values(
                    'ActiveInactive'
                ).annotate(
                    label=Coalesce('ActiveInactive', Value('Unknown')),
                    value=Count('FrameworkId')
                )
            elif base_model == 'policy':
                # For policies, use ActiveInactive field
                queryset = queryset.values(
                    'ActiveInactive'
                ).annotate(
                    label=Coalesce('ActiveInactive', Value('Unknown')),
                    value=Count('PolicyId')
                )
            else:
                # For subpolicies, use parent policy's ActiveInactive status
                queryset = queryset.values(
                    'PolicyId__ActiveInactive'  # Get ActiveInactive from parent policy
                ).annotate(
                    label=Coalesce('PolicyId__ActiveInactive', Value('Unknown')),
                    value=Count('SubPolicyId')
                )
        elif y_axis == 'framework_policies':
            if base_model == 'framework':
                queryset = queryset.annotate(
                    value=Count('policy')
                )
            else:
                queryset = queryset.annotate(value=Value(0))

        elif y_axis == 'createdByDate':
            # Handle CreatedByDate aggregation based on X-axis selection
            if x_axis == 'framework':
                queryset = queryset.values(
                    'CreatedByDate'
                ).annotate(
                    label=F('CreatedByDate'),
                    value=Count('FrameworkId')
                ).order_by('CreatedByDate')
            elif x_axis == 'policy':
                queryset = queryset.values(
                    'CreatedByDate'
                ).annotate(
                    label=F('CreatedByDate'),
                    value=Count('PolicyId')
                ).order_by('CreatedByDate')
            elif x_axis == 'subpolicy':
                queryset = queryset.values(
                    'CreatedByDate'
                ).annotate(
                    label=F('CreatedByDate'),
                    value=Count('SubPolicyId')
                ).order_by('CreatedByDate')

            # Format the dates for display
            data = list(queryset)
            for item in data:
                if item['label']:
                    # Convert date to string in YYYY-MM-DD format
                    item['label'] = item['label'].strftime('%Y-%m-%d')
            return Response(data)
        elif y_axis == 'department':
            # Handle Department aggregation based on X-axis selection
            if x_axis == 'framework':
                # For frameworks, get departments through policy relationship
                base_queryset = Framework.objects.values(
                    'FrameworkId', 'FrameworkName'
                ).annotate(
                    policy_count=Count('policy')
                ).filter(policy_count__gt=0)

                # Get all policies for these frameworks
                framework_policies = Policy.objects.filter(
                    FrameworkId__in=[f['FrameworkId'] for f in base_queryset]
                ).values('FrameworkId', 'Department')

                # Process departments and count frameworks
                department_counts = {}
                framework_departments = {}  # Track which departments each framework has been counted for

                for policy in framework_policies:
                    framework_id = policy['FrameworkId']
                    dept_str = policy['Department']
                    
                    if not dept_str:
                        continue

                    # Initialize set for this framework if not exists
                    if framework_id not in framework_departments:
                        framework_departments[framework_id] = set()

                    # Split departments and process each
                    departments = [d.strip().upper() for d in dept_str.split(',')]
                    for dept in departments:
                        if dept and dept not in framework_departments[framework_id]:
                            # Count framework for this department only once
                            department_counts[dept] = department_counts.get(dept, 0) + 1
                            framework_departments[framework_id].add(dept)

            elif x_axis == 'policy':
                # For policies, use Department field directly
                base_queryset = queryset.values(
                    'PolicyId',
                    'Department'
                )
            elif x_axis == 'subpolicy':
                # For subpolicies, get department through policy relationship
                base_queryset = queryset.values(
                    'SubPolicyId',
                    'PolicyId__Department'
                )

            if x_axis != 'framework':
                # Process departments and split comma-separated values
                department_counts = {}
                for item in base_queryset:
                    # Get the department field based on the model type
                    dept_field = (
                        item.get('Department') or 
                        item.get('policy__Department') or 
                        item.get('PolicyId__Department')
                    )
                    
                    if dept_field:
                        # Split departments by comma and strip whitespace
                        departments = [d.strip().upper() for d in dept_field.split(',')]
                        for dept in departments:
                            if dept:  # Only count non-empty departments
                                department_counts[dept] = department_counts.get(dept, 0) + 1

            # Convert to list format expected by frontend
            data = [
                {
                    'label': f"{dept.title()} ({count} frameworks)" if x_axis == 'framework' else f"{dept.title()} ({count} items)",
                    'value': count
                }
                for dept, count in department_counts.items()
            ]

            # Sort by count (descending) then by department name
            data.sort(key=lambda x: (-x['value'], x['label']))

            # Add unassigned if no departments found
            if not data:
                label = 'Unassigned (0 frameworks)' if x_axis == 'framework' else 'Unassigned (0 items)'
                data.append({
                    'label': label,
                    'value': 0
                })

            return Response(data)
        elif y_axis == 'createdByName':
            # Handle CreatedByName aggregation based on X-axis selection
            if x_axis == 'framework':
                queryset = queryset.values(
                    'CreatedByName'
                ).annotate(
                    label=F('CreatedByName'),  # Use the actual CreatedByName value
                    value=Count('FrameworkId', distinct=True)  # Count unique frameworks
                ).order_by('-value', 'CreatedByName')  # Order by count desc, then name
            elif x_axis == 'policy':
                queryset = queryset.values(
                    'CreatedByName'
                ).annotate(
                    label=F('CreatedByName'),  # Use the actual CreatedByName value
                    value=Count('PolicyId', distinct=True)  # Count unique policies
                ).order_by('-value', 'CreatedByName')  # Order by count desc, then name
            elif x_axis == 'subpolicy':
                queryset = queryset.values(
                    'CreatedByName'
                ).annotate(
                    label=F('CreatedByName'),  # Use the actual CreatedByName value
                    value=Count('SubPolicyId', distinct=True)  # Count unique subpolicies
                ).order_by('-value', 'CreatedByName')  # Order by count desc, then name

            # Add creator label for clarity
            data = list(queryset)
            for item in data:
                item['label'] = f"{item['label']} ({item['value']} items)"
            return Response(data)
        elif y_axis == 'status':
            if base_model == 'framework':
                # For frameworks, directly use the Status field
                queryset = queryset.values(
                    'Status'
                ).annotate(
                    label=Coalesce('Status', Value('Unknown')),
                    value=Count('FrameworkId')
                )
            elif base_model == 'policy':
                # For policies, use their Status field
                queryset = queryset.values(
                    'Status'
                ).annotate(
                    label=Coalesce('Status', Value('Unknown')),
                    value=Count('PolicyId')
                )
            else:
                # For subpolicies, use their Status field
                queryset = queryset.values(
                    'Status'
                ).annotate(
                    label=Coalesce('Status', Value('Unknown')),
                    value=Count('SubPolicyId')
                )
        elif y_axis == 'category':
            if base_model == 'framework':
                # For frameworks, directly use the Category field
                queryset = queryset.values(
                    'Category'
                ).annotate(
                    label=Coalesce('Category', Value('Uncategorized')),
                    value=Count('FrameworkId')
                )
            elif base_model == 'policy':
                # For policies, get category through framework relationship
                queryset = queryset.values(
                    'FrameworkId__Category'
                ).annotate(
                    label=Coalesce('FrameworkId__Category', Value('Uncategorized')),
                    value=Count('PolicyId')
                )
            else:
                # For subpolicies, get category through policy->framework relationship
                queryset = queryset.values(
                    'PolicyId__FrameworkId__Category'
                ).annotate(
                    label=Coalesce('PolicyId__FrameworkId__Category', Value('Uncategorized')),
                    value=Count('SubPolicyId')
                )
        
        data = list(queryset)
        
        # Format dates for time-based analysis
        if x_axis == 'time':
            for item in data:
                date_value = item.get(date_field)
                item['label'] = date_value.strftime('%Y-%m-%d') if date_value else None
        
        return Response(data)
    except Exception as e:
        return Response(
            {'error': str(e)}, 
            status=500
        )

@api_view(['GET'])
@permission_classes([AllowAny])
def get_policy_kpis(request):
    try:
        # Get total policies count
        total_policies = Policy.objects.count()
        
        # Get active policies count
        active_policies = Policy.objects.filter(
            ActiveInactive='Active'
        ).count()

        # Get total users count for acknowledgement rate calculation
        total_users = Users.objects.count()

        # Get top 5 policies by acknowledgement rate
        policies = Policy.objects.filter(ActiveInactive='Active').annotate(
            acknowledgement_rate=Case(
                When(AcknowledgementCount__gt=0, 
                     then=ExpressionWrapper(
                         F('AcknowledgementCount') * 100.0 / total_users,
                         output_field=FloatField()
                     )),
                default=Value(0.0),
                output_field=FloatField(),
            )
        ).order_by('-acknowledgement_rate')[:5]

        top_acknowledged_policies = [
            {
                'policy_id': policy.PolicyId,
                'policy_name': policy.PolicyName,
                'acknowledged_count': policy.AcknowledgementCount,
                'total_users': total_users,
                'acknowledgement_rate': round(float(policy.acknowledgement_rate), 1)
            }
            for policy in policies
        ]

        # Get historical active policy counts for the last 12 months
        twelve_months_ago = datetime.date.today() - timedelta(days=365)
        monthly_counts = []
        
        # Get all policies with their creation dates
        policies = Policy.objects.filter(
            CreatedByDate__gte=twelve_months_ago
        ).values('CreatedByDate', 'ActiveInactive')
        
        # Group by month and count active policies
        month_data = {}
        current_date = datetime.date.today()
        
        # Initialize all months with 0
        for i in range(12):
            month_date = current_date - timedelta(days=30 * i)
            month_key = month_date.strftime('%Y-%m')
            month_data[month_key] = 0
        
        # Count active policies for each month
        for policy in policies:
            month_key = policy['CreatedByDate'].strftime('%Y-%m')
            if month_key in month_data and policy['ActiveInactive'] == 'Active':
                month_data[month_key] += 1
        
        # Convert to sorted list for last 12 months
        monthly_counts = [
            {
                'month': k,
                'count': v
            }
            for k, v in sorted(month_data.items(), reverse=True)
        ][:12]
        
        # Calculate revision metrics
        three_months_ago = datetime.date.today() - timedelta(days=90)
        
        # Get all policy versions with previous version links
        policy_versions = PolicyVersion.objects.filter(
            PreviousVersionId__isnull=False  # Has a previous version
        ).select_related('PolicyId')
        
        # Track revised policies and their revision counts
        revision_counts = {}  # Dictionary to track revisions per PreviousVersionId
        revised_policies_set = set()
        total_revisions = 0
        
        # Process each version that has a previous version
        for version in policy_versions:
            policy_id = version.PolicyId.PolicyId
            prev_version_id = version.PreviousVersionId
            
            # Count this as a revision
            revised_policies_set.add(policy_id)
            total_revisions += 1
            
            # Track revisions per previous version
            if prev_version_id not in revision_counts:
                revision_counts[prev_version_id] = 1
            else:
                revision_counts[prev_version_id] += 1
        
        # Calculate policies with multiple revisions
        multiple_revisions_count = sum(1 for count in revision_counts.values() if count > 1)
        
        # Calculate final metrics
        revised_policies = len(revised_policies_set)
        
        # Calculate revision rate using total policies as denominator
        revision_rate = 0
        if total_policies > 0:  # Avoid division by zero
            revision_rate = (revised_policies / total_policies) * 100
            revision_rate = min(revision_rate, 100)  # Cap at 100%

        # Calculate policy coverage by department
        departments = Policy.objects.values_list('Department', flat=True).distinct()
        department_coverage = []
        
        for dept in departments:
            if dept:  # Skip empty department values
                dept_policies = Policy.objects.filter(Department=dept)
                total_dept_policies = dept_policies.count()
                if total_dept_policies > 0:
                    avg_coverage = dept_policies.aggregate(
                        avg_coverage=Coalesce(Avg('CoverageRate'), 0.0)
                    )['avg_coverage']
                    
                    department_coverage.append({
                        'department': dept,
                        'coverage_rate': round(float(avg_coverage), 2),
                        'total_policies': total_dept_policies
                    })
        
        # Sort departments by coverage rate in descending order
        department_coverage.sort(key=lambda x: x['coverage_rate'], reverse=True)
        
        # Calculate overall average coverage rate
        overall_coverage = Policy.objects.aggregate(
            avg_coverage=Coalesce(Avg('CoverageRate'), 0.0)
        )['avg_coverage']
        
        return Response({
            'total_policies': total_policies,
            'active_policies': active_policies,
            'active_policies_trend': monthly_counts,  # Add historical trend data
            'revision_rate': round(revision_rate, 2),
            'revised_policies': revised_policies,
            'total_revisions': total_revisions,
            'policies_with_multiple_revisions': multiple_revisions_count,
            'measurement_period': '3 months',
            'coverage_metrics': {
                'overall_coverage_rate': round(float(overall_coverage), 2),
                'department_coverage': department_coverage
            },
            'top_acknowledged_policies': top_acknowledged_policies
        })
    except Exception as e:
        print(f"Error in get_policy_kpis: {str(e)}")
        return Response({
            'error': 'Error fetching policy KPIs',
            'details': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['POST'])
@permission_classes([AllowAny])
def acknowledge_policy(request, policy_id):
    try:
        policy = Policy.objects.get(PolicyId=policy_id)
        user_id = 3  # Hardcoded for testing as requested
        
        # Initialize acknowledged_users list if None
        acknowledged_users = policy.AcknowledgedUserIds if policy.AcknowledgedUserIds is not None else []
        
        # Check if user already acknowledged
        if user_id not in acknowledged_users:
            # Add user to acknowledged list
            acknowledged_users.append(user_id)
            policy.AcknowledgedUserIds = acknowledged_users
            policy.AcknowledgementCount = len(acknowledged_users)
            policy.save()
            
            return Response({
                'message': 'Policy acknowledged successfully',
                'acknowledged_users': policy.AcknowledgedUserIds,
                'acknowledgement_count': policy.AcknowledgementCount
            })
        else:
            return Response({
                'message': 'Policy already acknowledged by this user',
                'acknowledged_users': policy.AcknowledgedUserIds,
                'acknowledgement_count': policy.AcknowledgementCount
            })

    except Policy.DoesNotExist:
        return Response({
            'error': 'Policy not found'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        print(f"Error in acknowledge_policy: {str(e)}")  # Add logging
        return Response({
            'error': 'Error acknowledging policy',
            'details': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
@permission_classes([AllowAny])
def login_user(request):
    """
    Authenticate a user with Django session and return a success response with user data
    """
    try:
        from django.contrib.auth import authenticate, login
        from django.contrib.auth.models import User as DjangoUser
        from django.contrib.sessions.models import Session
        
        data = request.data
        username = data.get('username')
        password = data.get('password')
        
        if not username or not password:
            return Response({
                'success': False,
                'message': 'Username and password are required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Log the login attempt
        logger.info(f"Login attempt for user: {username}")
        
        # Find the user in the GRC Users table
        try:
            grc_user = Users.objects.get(UserName=username)
            
            # Validate password (in production, use proper hashing)
            if grc_user.Password == password:
                # Create or get Django User for session management
                django_user, created = DjangoUser.objects.get_or_create(
                    username=username,
                    defaults={
                        'email': grc_user.Email,
                        'first_name': grc_user.UserName,
                        'is_active': True,
                        'is_staff': False,
                        'is_superuser': username == 'admin.grc'
                    }
                )
                
                # If user was created, set a default password
                if created:
                    django_user.set_password(password)
                    django_user.save()
                
                # Authenticate and login with Django's session system
                # First try to authenticate
                auth_user = authenticate(request, username=username, password=password)
                if not auth_user:
                    # If authentication fails, update the password and try again
                    django_user.set_password(password)
                    django_user.save()
                    auth_user = authenticate(request, username=username, password=password)
                
                if auth_user:
                    # Login the user (creates session)
                    login(request, auth_user)
                    
                    # Store GRC user ID in session for RBAC
                    request.session['user_id'] = grc_user.UserId  # Changed from grc_user_id to user_id for RBACUtils
                    request.session['grc_user_id'] = grc_user.UserId  # Keep both for backwards compatibility
                    request.session['grc_username'] = grc_user.UserName
                    request.session.save()
                    
                    # Log user permissions on login using RBAC system
                    try:
                        from .rbac.utils import RBACUtils
                        RBACUtils.log_user_login_permissions(grc_user.UserId)
                    except Exception as rbac_error:
                        logger.warning(f"Could not log RBAC permissions for user {grc_user.UserId}: {rbac_error}")
                    
                    logger.info(f"Successful login for user: {username}")
                    
                    return Response({
                        'success': True,
                        'message': 'Login successful',
                        'user': {
                            'id': grc_user.UserId,
                            'username': grc_user.UserName,
                            'email': grc_user.Email,
                            'django_user_id': django_user.id,
                            'session_key': request.session.session_key
                        }
                    })
                else:
                    logger.error(f"Django authentication failed for user: {username}")
                    return Response({
                        'success': False,
                        'message': 'Authentication system error'
                    }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
                    
            else:
                # Invalid password
                logger.warning(f"Invalid password for user: {username}")
                return Response({
                    'success': False,
                    'message': 'Invalid credentials'
                }, status=status.HTTP_401_UNAUTHORIZED)
                
        except Users.DoesNotExist:
            # User not found
            logger.warning(f"User not found: {username}")
            return Response({
                'success': False,
                'message': 'Invalid credentials'
            }, status=status.HTTP_401_UNAUTHORIZED)
            
    except Exception as e:
        # Log the error
        logger.error(f"Login error: {str(e)}")
        import traceback
        logger.error(f"Login traceback: {traceback.format_exc()}")
        return Response({
            'success': False,
            'message': 'An error occurred during login'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['POST'])
def logout_user(request):
    """
    Log out a user (client-side only in this implementation)
    """
    return Response({
        'success': True,
        'message': 'Logged out successfully'
    })

@api_view(['POST'])
@permission_classes([AllowAny])
def register_user(request):
    """
    Register a new user
    """
    try:
        data = request.data
        username = data.get('username')
        password = data.get('password')
        email = data.get('email')
        
        if not username or not password or not email:
            return Response({
                'success': False,
                'message': 'Username, password, and email are required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Check if user already exists
        if Users.objects.filter(UserName=username).exists():
            return Response({
                'success': False,
                'message': 'Username already exists'
            }, status=status.HTTP_400_BAD_REQUEST)
            
        if Users.objects.filter(email=email).exists():
            return Response({
                'success': False,
                'message': 'Email already exists'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Create new user
        user = Users.objects.create(
            UserName=username,
            Password=password,  # In production, this should be hashed
            email=email
        )
        
        return Response({
            'success': True,
            'message': 'User registered successfully',
            'user': {
                'id': user.UserId,
                'username': user.UserName,
                'email': user.email
            }
        }, status=status.HTTP_201_CREATED)
        
    except Exception as e:
        logger.error(f"Registration error: {str(e)}")
        return Response({
            'success': False,
            'message': 'An error occurred during registration'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([AllowAny])
def test_connection(request):
    """
    Test API connection
    """
    return Response({
        'success': True,
        'message': 'API connection successful'
    })

@api_view(['GET'])
@permission_classes([AllowAny])
def get_user_details_by_id(request, user_id):
    """
    Get user details from Users table by user ID
    """
    try:
        # Get user from Users table
        user = Users.objects.get(UserId=user_id)
        
        # Get RBAC info if available
        rbac_info = None
        try:
            from .rbac.utils import RBACUtils
            rbac_info = RBACUtils.get_user_rbac_info(user_id)
        except Exception as rbac_error:
            logger.warning(f"Could not get RBAC info for user {user_id}: {rbac_error}")
        
        user_details = {
            'user_id': user.UserId,
            'username': user.UserName,
            'email': user.email,
            'created_at': user.CreatedAt,
            'updated_at': user.UpdatedAt,
            'rbac_info': rbac_info
        }
        
        logger.info(f"✅ Retrieved user details for ID {user_id}: {user.UserName}")
        
        return Response({
            'success': True,
            'user': user_details,
            'message': f'User details retrieved for {user.UserName}'
        })
        
    except Users.DoesNotExist:
        logger.warning(f"❌ User with ID {user_id} not found")
        return Response({
            'success': False,
            'message': f'User with ID {user_id} not found'
        }, status=404)
        
    except Exception as e:
        logger.error(f"❌ Error retrieving user details for ID {user_id}: {str(e)}")
        return Response({
            'success': False,
            'message': 'Error retrieving user details',
            'error': str(e)
        }, status=500)

@api_view(['POST'])
@permission_classes([AllowAny])
def save_user_session(request):
    """
    Save user session information manually for testing
    """
    try:
        from django.contrib.auth import authenticate, login
        from django.contrib.auth.models import User as DjangoUser
        
        data = request.data
        user_id = data.get('user_id')
        
        if not user_id:
            return Response({
                'success': False,
                'message': 'user_id is required'
            }, status=400)
        
        # Get user from Users table
        try:
            grc_user = Users.objects.get(UserId=user_id)
        except Users.DoesNotExist:
            return Response({
                'success': False,
                'message': f'GRC User with ID {user_id} not found'
            }, status=404)
        
        # Create or get Django User for session management
        django_user, created = DjangoUser.objects.get_or_create(
            username=grc_user.UserName,
            defaults={
                'email': grc_user.Email,
                'first_name': grc_user.UserName,
                'is_active': True,
                'is_staff': False,
                'is_superuser': grc_user.UserName == 'admin.grc'
            }
        )
        
        # Set a default password if user was created
        if created:
            django_user.set_password('default_password')
            django_user.save()
        
        # Login the user (creates session)
        login(request, django_user)
        
        # Store GRC user ID in session for RBAC
        request.session['user_id'] = grc_user.UserId  # Changed from grc_user_id to user_id for RBACUtils
        request.session['grc_user_id'] = grc_user.UserId  # Keep both for backwards compatibility
        request.session['grc_username'] = grc_user.UserName
        request.session['grc_email'] = grc_user.email
        request.session.save()
        
        # Log user permissions using RBAC system
        try:
            from .rbac.utils import RBACUtils
            RBACUtils.log_user_login_permissions(grc_user.UserId)
        except Exception as rbac_error:
            logger.warning(f"Could not log RBAC permissions for user {grc_user.UserId}: {rbac_error}")
        
        logger.info(f"✅ Session created for user: {grc_user.UserName} (ID: {grc_user.UserId})")
        
        return Response({
            'success': True,
            'message': f'Session created for user {grc_user.UserName}',
            'user': {
                'id': grc_user.UserId,
                'username': grc_user.UserName,
                'email': grc_user.Email,
                'django_user_id': django_user.id,
                'session_key': request.session.session_key
            },
            'session_data': dict(request.session)
        })
        
    except Exception as e:
        logger.error(f"❌ Error saving user session: {str(e)}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return Response({
            'success': False,
            'message': 'Error saving user session',
            'error': str(e)
        }, status=500)


